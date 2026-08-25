from __future__ import annotations

import tempfile
import unittest
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from src.attack_integrity import _equivalent_value, validate_attack_integrity
from src.operators import ROBUSTNESS_OPERATORS
from src.workspace import AttackWorkspace


def feature_payload(file_name: str, sheet: str | None = None):
    pack = {
        "file": file_name, "header_row": 1, "target_fields": ["value"], "misuse_answer": "different total",
        "added_features": [
            {"name": "value_forecast", "source_field": "value", "noise_subtype": "forecast_feature_pack", "exclusion_reason": "question requests observed value"},
            {"name": "value_rank", "source_field": "value", "noise_subtype": "ranking_feature_pack", "exclusion_reason": "rank is not raw value"},
        ],
    }
    if sheet is not None:
        pack["sheet"] = sheet
    return {"transformation_record": {"semantic_contract": {"feature_pack": pack}}}


def non_observation_payload(file_name: str, sheet: str | None = None):
    row = {
        "file": file_name, "row": 12, "marker_field": "record_type", "marker_value": "simulated",
        "noise_subtype": "simulated_row", "exclusion_reason": "record_type identifies synthetic control", "misuse_answer": "different total",
    }
    if sheet is not None:
        row["sheet"] = sheet
    return {"transformation_record": {"semantic_contract": {
        "marker_column": {"name": "record_type", "observation_value": "observation"},
        "non_observation_rows": [row],
    }}}


class AttackIntegrityTests(unittest.TestCase):
    def _excel_workspace(self, root: Path, attack: str) -> AttackWorkspace:
        source = root / "data.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["id", "category", "value"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for index in range(1, 11):
            ws.append([index, f"{index:02d}", float(index * 10)])
            ws.cell(index + 1, 3).number_format = "0.00"
        wb.save(source)
        wb.close()
        workspace = AttackWorkspace(root / "outputs", f"sample__{attack}__001", "sample", attack)
        workspace.prepare([source])
        return workspace

    def _report(self, workspace, payload):
        return validate_attack_integrity(workspace, workspace.attack_type, payload, {})

    def test_equivalent_value_normalizes_dates_and_numbers(self):
        from datetime import datetime
        self.assertTrue(_equivalent_value("2000-01-01", datetime(2000, 1, 1)))
        self.assertTrue(_equivalent_value("1,234.00", 1234))
        self.assertTrue(_equivalent_value("50%", 0.5))
        self.assertFalse(_equivalent_value("01/02/2000", "2000-01-02"))

    def test_excel_feature_pack_passes_and_rejects_original_change(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            workspace = self._excel_workspace(root, "decoy_feature_pack_injection")
            target = workspace.final / "data.xlsx"
            wb = load_workbook(target)
            ws = wb["Data"]
            for col, name, multiplier in ((4, "value_forecast", 1.1), (5, "value_rank", 0.1)):
                ws.cell(1, col, name)._style = copy(ws["C1"]._style)
                for row in range(2, 12):
                    ws.cell(row, col, ws.cell(row, 3).value * multiplier).number_format = "0.00"
            wb.save(target); wb.close()
            self.assertTrue(self._report(workspace, feature_payload("data.xlsx", "Data")).passed)

            workspace.prepare([root / "data.xlsx"])
            target = workspace.final / "data.xlsx"
            wb = load_workbook(target)
            ws = wb["Data"]
            ws["C2"] = 999
            for col, name in ((4, "value_forecast"), (5, "value_rank")):
                ws.cell(1, col, name)._style = copy(ws["C1"]._style)
                for row in range(2, 12):
                    ws.cell(row, col, float(row))
            wb.save(target); wb.close()
            report = self._report(workspace, feature_payload("data.xlsx", "Data"))
            self.assertFalse(report.passed)
            self.assertIn("original_records_changed", report.error_message())

    def test_excel_non_observation_rows_pass_and_reject_bad_marker(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            workspace = self._excel_workspace(root, "non_observation_row_injection")
            target = workspace.final / "data.xlsx"
            wb = load_workbook(target)
            ws = wb["Data"]
            ws["D1"] = "record_type"; ws["D1"]._style = copy(ws["C1"]._style)
            for row in range(2, 12): ws.cell(row, 4, "observation")
            ws.append([101, "11", 999.0, "simulated"])
            wb.save(target); wb.close()
            self.assertTrue(self._report(workspace, non_observation_payload("data.xlsx", "Data")).passed)

            workspace.prepare([root / "data.xlsx"])
            target = workspace.final / "data.xlsx"
            wb = load_workbook(target); ws = wb["Data"]
            ws["D1"] = "record_type"; ws["D1"]._style = copy(ws["C1"]._style)
            for row in range(2, 12): ws.cell(row, 4, "observation")
            ws.append([101, "11", 999.0, "observation"])
            wb.save(target); wb.close()
            report = self._report(workspace, non_observation_payload("data.xlsx", "Data"))
            self.assertFalse(report.passed)
            self.assertIn("non_observation_marker_invalid", report.error_message())

    def test_csv_l4_contracts_preserve_lexical_values(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "data.csv"
            source.write_text("id,category,value\n01,a,10\n02,b,20\n03,c,30\n", encoding="utf-8-sig")
            feature = AttackWorkspace(root / "outputs", "feature__001", "feature", "decoy_feature_pack_injection")
            feature.prepare([source])
            (feature.final / "data.csv").write_text("id,category,value,value_forecast,value_rank\n01,a,10,11,1\n02,b,20,22,2\n03,c,30,33,3\n", encoding="utf-8-sig")
            self.assertTrue(self._report(feature, feature_payload("data.csv")).passed)

            rows = AttackWorkspace(root / "outputs", "rows__001", "rows", "non_observation_row_injection")
            rows.prepare([source])
            (rows.final / "data.csv").write_text("id,category,value,record_type\n01,a,10,observation\n02,b,20,observation\n03,c,30,observation\n04,d,999,simulated\n", encoding="utf-8-sig")
            payload = non_observation_payload("data.csv")
            payload["transformation_record"]["semantic_contract"]["non_observation_rows"][0]["row"] = 5
            self.assertTrue(self._report(rows, payload).passed)
            (rows.final / "data.csv").write_text("id,category,value,record_type\n01,a,10,observation\n02,b,20,observation\n03,c,30,observation\n04,d,999,observation\n", encoding="utf-8-sig")
            self.assertFalse(self._report(rows, payload).passed)

    def test_row_shuffle_still_preserves_typed_rows(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            workspace = self._excel_workspace(root, "row_order_shuffle")
            target = workspace.final / "data.xlsx"
            wb = load_workbook(target); ws = wb["Data"]
            values = [[cell.value for cell in ws[row]] for row in range(2, 12)]
            for row_index, row in enumerate(reversed(values), start=2):
                for col, value in enumerate(row, start=1): ws.cell(row_index, col, value)
            wb.save(target); wb.close()
            self.assertTrue(self._report(workspace, {"transformation_record": {"parameters": {}}}).passed)

    def test_column_shuffle_ignores_trailing_formatted_empty_columns(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            workspace = self._excel_workspace(root, "column_order_shuffle")
            source = root / "data.xlsx"
            # Excel's used range may include a styled blank column. It is not a
            # table field and must remain fixed while the active header region moves.
            wb = load_workbook(source)
            ws = wb["Data"]
            ws["D1"]._style = copy(ws["C1"]._style)
            wb.save(source); wb.close()
            workspace.prepare([source])
            target = workspace.final / "data.xlsx"
            wb = load_workbook(target); ws = wb["Data"]
            columns = [
                [(copy(cell._style), cell.value, cell.number_format) for cell in column]
                for column in ws.iter_cols(min_col=1, max_col=3)
            ]
            for destination, source_col in enumerate((3, 1, 2), start=1):
                for row, (style, value, number_format) in enumerate(columns[source_col - 1], start=1):
                    cell = ws.cell(row, destination)
                    cell._style = copy(style)
                    cell.value = value
                    cell.number_format = number_format
            wb.save(target); wb.close()
            self.assertTrue(self._report(workspace, {"transformation_record": {"parameters": {}}}).passed)

    def test_all_registered_operators_have_contract_coverage(self):
        expected = {
            "row_order_shuffle", "column_order_shuffle", "header_synonym_substitution", "semantic_distractor_column",
            "equivalent_value_reencoding", "unit_scale_conversion", "csv_wide_long_reshape", "csv_relational_decomposition",
            "excel_hierarchical_header_relayout", "excel_cross_sheet_relayout", "decoy_feature_pack_injection", "non_observation_row_injection",
        }
        self.assertEqual(expected, set(ROBUSTNESS_OPERATORS))


if __name__ == "__main__":
    unittest.main()
