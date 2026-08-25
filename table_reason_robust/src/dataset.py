from __future__ import annotations

import json
import os
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional


def load_jsonl(path: Path) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            if line.strip():
                rows.append(json.loads(line))
    return rows


def build_dabench_question(row: Dict[str, Any]) -> str:
    """Preserve DABENCH's task constraints as part of the executable question."""
    parts = [str(row.get("question") or "").strip()]
    constraints = str(row.get("constraints") or "").strip()
    required_format = str(row.get("format") or "").strip()
    if constraints:
        parts.append(f"Constraints:\n{constraints}")
    if required_format:
        parts.append(f"Required output format:\n{required_format}")
    return "\n\n".join(part for part in parts if part)


def load_dabench_rows(
    questions_path: Path,
    labels_path: Path,
    table_root: Path,
    selected_ids: Optional[Iterable[str]] = None,
) -> List[Dict[str, Any]]:
    """Merge official DABENCH files into the runner's common row contract.

    This intentionally has no missing-table fallback.  ``selected_ids`` exists so
    a bounded smoke run validates exactly the requested source subset; a normal
    full run passes ``None`` and therefore validates the entire benchmark.
    """
    for label, path in (("questions_path", questions_path), ("labels_path", labels_path)):
        if not path.is_file():
            raise FileNotFoundError(f"DABENCH {label} does not exist: {path}")
    if not table_root.is_dir():
        raise FileNotFoundError(f"DABENCH table_root does not exist: {table_root}")

    questions = load_jsonl(questions_path)
    labels = load_jsonl(labels_path)
    labels_by_id: Dict[str, Dict[str, Any]] = {}
    for label in labels:
        original_id = str(label.get("id", ""))
        if not original_id:
            raise RuntimeError("DABENCH label is missing id")
        if original_id in labels_by_id:
            raise RuntimeError(f"DABENCH labels contain duplicate id: {original_id}")
        labels_by_id[original_id] = label

    requested = set(str(value) for value in selected_ids) if selected_ids else None
    seen_question_ids = set()
    rows: List[Dict[str, Any]] = []
    for question in questions:
        original_id = str(question.get("id", ""))
        if not original_id:
            raise RuntimeError("DABENCH question is missing id")
        if original_id in seen_question_ids:
            raise RuntimeError(f"DABENCH questions contain duplicate id: {original_id}")
        seen_question_ids.add(original_id)
        if requested is not None and original_id not in requested and f"DA_{original_id}" not in requested:
            continue
        label = labels_by_id.get(original_id)
        if label is None:
            raise RuntimeError(f"DABENCH question id has no matching label: {original_id}")
        file_name = str(question.get("file_name") or "").strip()
        if not file_name:
            raise RuntimeError(f"DABENCH question {original_id} is missing file_name")
        file_path = table_root / file_name
        if file_path.suffix.lower() != ".csv":
            raise RuntimeError(f"DABENCH question {original_id} does not reference a CSV: {file_name}")
        if not file_path.is_file():
            raise FileNotFoundError(f"DABENCH question {original_id} references missing CSV: {file_path}")
        common_answers = label.get("common_answers")
        if common_answers is None:
            raise RuntimeError(f"DABENCH label {original_id} is missing common_answers")
        rows.append({
            "id": f"DA_{original_id}",
            "question": build_dabench_question(question),
            "input_file": file_name,
            "reference": json.dumps({"common_answers": common_answers}, ensure_ascii=False),
            "output_type": "data_analysis",
            "metadata": {
                "source": "DABENCH",
                "original_id": question.get("id"),
                "concepts": question.get("concepts", []),
                "level": question.get("level"),
                "original_question": question.get("question"),
                "constraints": question.get("constraints"),
                "required_format": question.get("format"),
            },
        })
    if requested is not None:
        found = {str(row["id"]) for row in rows}
        missing = sorted(
            value for value in requested
            if value if (value if value.startswith("DA_") else f"DA_{value}") not in found
        )
        if missing:
            raise RuntimeError(f"DABENCH selected sample ids not found: {missing}")
    if set(labels_by_id) != seen_question_ids:
        missing_questions = sorted(set(labels_by_id) - seen_question_ids)
        missing_labels = sorted(seen_question_ids - set(labels_by_id))
        raise RuntimeError(
            "DABENCH questions/labels are not one-to-one: "
            f"labels_without_questions={missing_questions[:5]}, questions_without_labels={missing_labels[:5]}"
        )
    return rows


def split_input_files(value: Any) -> List[str]:
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    return [part.strip() for part in re.split(r"[\n;；]+", text) if part.strip()]


def resolve_input_files(row: Dict[str, Any], data_root: Path) -> List[Path]:
    sample_id = str(row.get("id", ""))
    files = split_input_files(row.get("input_file", ""))
    resolved: List[Path] = []
    for name in files:
        path = Path(name)
        if path.is_absolute() and path.exists():
            resolved.append(path)
            continue
        candidates = [
            data_root / sample_id / name,
            data_root / name,
        ]
        for candidate in candidates:
            if candidate.exists():
                resolved.append(candidate)
                break
        else:
            resolved.append(data_root / sample_id / name)
    return resolved


def rows_to_limit(rows: List[Dict[str, Any]], limit: Optional[int]) -> List[Dict[str, Any]]:
    if limit is None:
        return rows
    return rows[: int(limit)]


def profile_files(file_paths: Iterable[Path], virtual_root: str = "/mnt/data") -> Dict[str, Any]:
    profiles: List[Dict[str, Any]] = []
    for path in file_paths:
        item: Dict[str, Any] = {
            "name": path.name,
            "virtual_path": f"{virtual_root}/{path.name}",
            "exists": path.exists(),
            "size_bytes": path.stat().st_size if path.exists() else None,
            "sheets": [],
        }
        if not path.exists():
            item["error"] = "file_not_found"
            profiles.append(item)
            continue
        suffix = path.suffix.lower()
        try:
            if suffix in {".xlsx", ".xlsm"}:
                item["sheets"] = _profile_excel(path)
            elif suffix in {".csv", ".tsv"}:
                item["sheets"] = [_profile_csv(path, sep="\t" if suffix == ".tsv" else ",")]
            elif suffix == ".xls":
                item["note"] = "unsupported_attack_target:.xls"
            else:
                item["note"] = f"unsupported_profile_type:{suffix}"
        except Exception as exc:
            item["error"] = f"profile_failed:{exc}"
        profiles.append(item)
    return {"files": profiles}


def _profile_excel(path: Path) -> List[Dict[str, Any]]:
    import pandas as pd

    sheets: List[Dict[str, Any]] = []
    excel = pd.ExcelFile(path)
    for sheet_name in excel.sheet_names[:10]:
        try:
            sample = pd.read_excel(path, sheet_name=sheet_name, nrows=5)
            row_count = _excel_row_count(path, sheet_name)
            profile = _frame_profile(sheet_name, sample, row_count)
            profile.update(_excel_structure(path, sheet_name))
            sheets.append(profile)
        except Exception as exc:
            sheets.append({"name": sheet_name, "error": str(exc)})
    return sheets


def _profile_csv(path: Path, sep: str) -> Dict[str, Any]:
    import pandas as pd

    sample = pd.read_csv(path, sep=sep, nrows=5)
    with path.open("r", encoding="utf-8", errors="ignore") as stream:
        row_count = sum(1 for _ in stream) - 1
    return _frame_profile("CSV", sample, max(row_count, 0))


def _frame_profile(name: str, df: Any, row_count: Optional[int]) -> Dict[str, Any]:
    sample_rows = json.loads(df.head(5).to_json(orient="records", force_ascii=False))
    return {
        "name": str(name),
        "columns": [str(col) for col in df.columns],
        "dtypes": {str(col): str(dtype) for col, dtype in df.dtypes.items()},
        "n_rows": row_count,
        "sample_rows": sample_rows,
    }


def _excel_row_count(path: Path, sheet_name: str) -> Optional[int]:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            return wb[sheet_name].max_row
        finally:
            wb.close()
    except Exception:
        return None


def _excel_structure(path: Path, sheet_name: str) -> Dict[str, Any]:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=False, data_only=False, keep_vba=path.suffix.lower() == ".xlsm")
        try:
            ws = wb[sheet_name]
            merged = [str(item) for item in list(ws.merged_cells.ranges)[:30]]
            top_rows: List[List[Any]] = []
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8), max_col=min(ws.max_column, 20)):
                top_rows.append([_json_safe(cell.value) for cell in row])
            formula_count = sum(
                1
                for row in ws.iter_rows()
                for cell in row
                if isinstance(cell.value, str) and cell.value.startswith("=")
            )
            return {
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "merged_ranges": merged,
                "formula_count": formula_count,
                "top_rows": top_rows,
            }
        finally:
            wb.close()
    except Exception as exc:
        return {"structure_error": str(exc)}


def _json_safe(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def compact_json(data: Any) -> str:
    return json.dumps(data, ensure_ascii=False, indent=2)


def project_root_from_file(file: str) -> Path:
    return Path(file).resolve().parents[1]
