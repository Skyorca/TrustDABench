from __future__ import annotations

import csv
import hashlib
import math
import re
import threading
from collections import Counter, OrderedDict
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple


INTEGRITY_SCHEMA_VERSION = 2
SHUFFLE_ATTACKS = {"row_order_shuffle", "column_order_shuffle"}
L4_FEATURE_PACK_ATTACK = "decoy_feature_pack_injection"
L4_NON_OBSERVATION_ATTACK = "non_observation_row_injection"
ROW_INJECTION_ATTACKS = {L4_NON_OBSERVATION_ATTACK}
RESHAPE_ATTACKS = {
    "csv_wide_long_reshape",
    "csv_relational_decomposition",
    "excel_hierarchical_header_relayout",
    "excel_cross_sheet_relayout",
}
_ORIGINAL_WORKBOOK_CACHE: "OrderedDict[Tuple[Any, ...], Any]" = OrderedDict()
_ORIGINAL_WORKBOOK_CACHE_LOCK = threading.Lock()
_ORIGINAL_WORKBOOK_CACHE_SIZE = 4


@dataclass
class IntegrityIssue:
    code: str
    message: str
    file: Optional[str] = None
    sheet: Optional[str] = None
    coordinate: Optional[str] = None
    expected: Optional[str] = None
    actual: Optional[str] = None


@dataclass
class IntegrityReport:
    attack_type: str
    passed: bool = True
    changed: bool = False
    checks: Dict[str, Any] = field(default_factory=dict)
    issues: List[IntegrityIssue] = field(default_factory=list)
    schema_version: int = INTEGRITY_SCHEMA_VERSION

    def add(
        self,
        code: str,
        message: str,
        *,
        file: Optional[str] = None,
        sheet: Optional[str] = None,
        coordinate: Optional[str] = None,
        expected: Any = None,
        actual: Any = None,
    ) -> None:
        self.passed = False
        self.issues.append(
            IntegrityIssue(
                code=code,
                message=message,
                file=file,
                sheet=sheet,
                coordinate=coordinate,
                expected=_short(expected),
                actual=_short(actual),
            )
        )

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)

    def error_message(self) -> Optional[str]:
        if self.passed:
            return None
        parts = []
        for issue in self.issues[:12]:
            location = "/".join(item for item in (issue.file, issue.sheet, issue.coordinate) if item)
            text = f"{issue.code}: {issue.message}"
            if location:
                text += f" [{location}]"
            if issue.expected is not None or issue.actual is not None:
                text += f" expected={issue.expected} actual={issue.actual}"
            parts.append(text)
        return "integrity check failed: " + "; ".join(parts)


def validate_attack_integrity(
    workspace: Any,
    attack_type: str,
    construct_payload: Optional[Dict[str, Any]] = None,
    selection: Optional[Dict[str, Any]] = None,
) -> IntegrityReport:
    payload = construct_payload or {}
    selection = selection or {}
    report = IntegrityReport(attack_type=attack_type)
    original_names = set(workspace.original_file_names())
    final_names = set(workspace.final_file_names())

    if attack_type in {"header_synonym_substitution", "semantic_distractor_column"}:
        _validate_l1_semantic_contract(attack_type, payload, selection, report)
    if attack_type in {L4_FEATURE_PACK_ATTACK, L4_NON_OBSERVATION_ATTACK}:
        _validate_l4_semantic_contract(attack_type, payload, report)

    allows_package_change = attack_type == "csv_relational_decomposition"
    if not allows_package_change and original_names != final_names:
        report.add(
            "file_package_changed",
            "attack must preserve the file package",
            expected=sorted(original_names),
            actual=sorted(final_names),
        )
        return report
    if attack_type == "csv_relational_decomposition":
        _validate_relational_package(workspace, report, payload)
        return _finish(report)

    for name in sorted(original_names & final_names):
        original = workspace.original / name
        attacked = workspace.final / name
        suffix = original.suffix.lower()
        if suffix in {".xlsx", ".xlsm"}:
            _validate_excel_pair(original, attacked, attack_type, payload, selection, report)
        elif suffix in {".csv", ".tsv"}:
            _validate_delimited_pair(
                original,
                attacked,
                attack_type,
                payload,
                selection,
                report,
                "\t" if suffix == ".tsv" else ",",
            )
        elif original.read_bytes() != attacked.read_bytes():
            report.add("unsupported_file_changed", "unsupported companion file changed", file=name)

    return _finish(report)


def _finish(report: IntegrityReport) -> IntegrityReport:
    if not report.changed:
        report.add("attack_not_changed", "no permitted content change was detected")
    report.checks["issue_count"] = len(report.issues)
    return report


def _validate_excel_pair(
    original_path: Path,
    attacked_path: Path,
    attack_type: str,
    payload: Dict[str, Any],
    selection: Dict[str, Any],
    report: IntegrityReport,
) -> None:
    from openpyxl import load_workbook

    keep_vba = original_path.suffix.lower() == ".xlsm"
    try:
        original_wb = _load_original_workbook_cached(original_path, keep_vba)
        attacked_wb = load_workbook(attacked_path, data_only=False, read_only=False, keep_vba=keep_vba)
    except Exception as exc:
        report.add("invalid_file", f"workbook cannot be opened: {exc}", file=original_path.name)
        return
    try:
        if attack_type == "excel_cross_sheet_relayout":
            _validate_excel_containers(original_wb, attacked_wb, attack_type, original_path.name, payload, report)
            return
        if original_wb.sheetnames != attacked_wb.sheetnames:
            report.add(
                "sheet_names_changed",
                "sheet names changed unexpectedly",
                file=original_path.name,
                expected=original_wb.sheetnames,
                actual=attacked_wb.sheetnames,
            )
            return
        for sheet_name in original_wb.sheetnames:
            left = original_wb[sheet_name]
            right = attacked_wb[sheet_name]
            _validate_sheet_structure(left, right, attack_type, original_path.name, report)
            if attack_type == "row_order_shuffle":
                _validate_row_shuffle(left, right, original_path.name, report)
            elif attack_type == "column_order_shuffle":
                _validate_column_shuffle(left, right, original_path.name, report)
            elif attack_type == "header_synonym_substitution":
                _validate_header_synonyms(left, right, original_path.name, payload, report)
            elif attack_type == "semantic_distractor_column":
                _validate_distractor_columns(left, right, original_path.name, payload, report)
            elif attack_type == L4_FEATURE_PACK_ATTACK:
                _validate_excel_decoy_feature_pack(left, right, original_path.name, payload, report)
            elif attack_type == "equivalent_value_reencoding":
                _validate_reencoding(left, right, original_path.name, report)
            elif attack_type == "unit_scale_conversion":
                _validate_unit_conversion(left, right, original_path.name, report)
            elif attack_type == L4_NON_OBSERVATION_ATTACK:
                _validate_excel_non_observation_rows(left, right, original_path.name, payload, report)
            elif attack_type == "excel_hierarchical_header_relayout":
                _validate_hierarchical_relayout(left, right, original_path.name, payload, report)
            else:
                _validate_preserved_or_changed(left, right, original_path.name, report)
    finally:
        attacked_wb.close()


def _validate_sheet_structure(left: Any, right: Any, attack_type: str, file_name: str, report: IntegrityReport) -> None:
    allow_shape_change = attack_type in {
        "semantic_distractor_column",
        "excel_hierarchical_header_relayout",
        L4_FEATURE_PACK_ATTACK,
        *ROW_INJECTION_ATTACKS,
    }
    if not allow_shape_change and (left.max_row, left.max_column) != (right.max_row, right.max_column):
        report.add(
            "sheet_shape_changed",
            "sheet shape changed unexpectedly",
            file=file_name,
            sheet=left.title,
            expected=(left.max_row, left.max_column),
            actual=(right.max_row, right.max_column),
        )
    if attack_type != "excel_hierarchical_header_relayout":
        left_merged = sorted(str(item) for item in left.merged_cells.ranges)
        right_merged = sorted(str(item) for item in right.merged_cells.ranges)
        if left_merged != right_merged:
            report.add(
                "merged_ranges_changed",
                "merged header/data ranges changed unexpectedly",
                file=file_name,
                sheet=left.title,
                expected=left_merged,
                actual=right_merged,
            )
    if attack_type not in {"semantic_distractor_column", "excel_hierarchical_header_relayout", *ROW_INJECTION_ATTACKS}:
        if _filter_token(left) != _filter_token(right):
            report.add("filter_changed", "filter range or criteria changed", file=file_name, sheet=left.title)
        if _table_tokens(left) != _table_tokens(right):
            report.add("table_range_changed", "Excel table names or ranges changed", file=file_name, sheet=left.title)
    if _hidden_rows(left) != _hidden_rows(right) and attack_type != "row_order_shuffle":
        report.add("hidden_rows_changed", "hidden row state changed", file=file_name, sheet=left.title)
    if _hidden_columns(left) != _hidden_columns(right) and attack_type != "column_order_shuffle":
        report.add("hidden_columns_changed", "hidden column state changed", file=file_name, sheet=left.title)
    if attack_type in {"equivalent_value_reencoding", "unit_scale_conversion", *ROW_INJECTION_ATTACKS}:
        for col in range(1, min(left.max_column, right.max_column) + 1):
            old = left.cell(1, col)
            new = right.cell(1, col)
            if (tuple(old._style) if old.has_style else ()) != (tuple(new._style) if new.has_style else ()):
                report.add(
                    "header_style_changed",
                    "header style changed outside a header-relayout attack",
                    file=file_name,
                    sheet=left.title,
                    coordinate=old.coordinate,
                )


def _validate_row_shuffle(left: Any, right: Any, file_name: str, report: IntegrityReport) -> None:
    if (left.max_row, left.max_column) != (right.max_row, right.max_column) or left.max_row < 2:
        return
    left_header = tuple(_header_token(cell) for cell in left[1])
    right_header = tuple(_header_token(cell) for cell in right[1])
    if left_header != right_header:
        report.add("header_changed", "header values/types/styles changed", file=file_name, sheet=left.title)
    left_rows = [tuple(_data_token(cell) for cell in row) for row in left.iter_rows(min_row=2)]
    right_rows = [tuple(_data_token(cell) for cell in row) for row in right.iter_rows(min_row=2)]
    if Counter(left_rows) != Counter(right_rows):
        report.add("data_row_multiset_changed", "typed/formatted data-row multiset changed", file=file_name, sheet=left.title)
    moved = sum(a != b for a, b in zip(left_rows, right_rows))
    ratio = moved / len(left_rows) if left_rows else 0.0
    report.checks[f"{file_name}/{left.title}/row_move_ratio"] = round(ratio, 6)
    if left_rows and ratio < 0.8:
        report.add("row_shuffle_too_weak", "fewer than 80% of rows changed position", actual=ratio)
    report.changed = report.changed or moved > 0


def _validate_column_shuffle(left: Any, right: Any, file_name: str, report: IntegrityReport) -> None:
    if (left.max_row, left.max_column) != (right.max_row, right.max_column) or left.max_column < 2:
        return
    left_width = _active_header_width(left)
    right_width = _active_header_width(right)
    if left_width != right_width or left_width < 2:
        report.add(
            "header_region_changed",
            "column shuffle changed the active single-row header region",
            file=file_name,
            sheet=left.title,
            expected=left_width,
            actual=right_width,
        )
        return
    # Workbooks frequently retain formatted, empty columns to the right of the
    # actual table. They are not table fields and must stay fixed, not make an
    # otherwise valid three-column table look like it has duplicate blank headers.
    for col in range(left_width + 1, left.max_column + 1):
        left_tokens = tuple(_column_token(left.cell(row, col), row == 1) for row in range(1, left.max_row + 1))
        right_tokens = tuple(_column_token(right.cell(row, col), row == 1) for row in range(1, right.max_row + 1))
        if left_tokens != right_tokens:
            report.add(
                "non_table_column_changed",
                "formatted empty column outside the active table changed",
                file=file_name,
                sheet=left.title,
                coordinate=left.cell(1, col).coordinate,
            )
    left_headers = [left.cell(1, col).value for col in range(1, left_width + 1)]
    right_headers = [right.cell(1, col).value for col in range(1, right_width + 1)]
    if len(set(left_headers)) != len(left_headers) or len(set(right_headers)) != len(right_headers):
        report.add("duplicate_headers", "column shuffle requires unique single-row headers", file=file_name, sheet=left.title)
        return
    if set(left_headers) != set(right_headers):
        report.add("header_set_changed", "header set changed", file=file_name, sheet=left.title)
        return
    for header in left_headers:
        old_col = left_headers.index(header) + 1
        new_col = right_headers.index(header) + 1
        left_tokens = tuple(_column_token(left.cell(row, old_col), row == 1) for row in range(1, left.max_row + 1))
        right_tokens = tuple(_column_token(right.cell(row, new_col), row == 1) for row in range(1, right.max_row + 1))
        if left_tokens != right_tokens:
            report.add(
                "column_content_changed",
                f"column content/type/format changed for {header!r}",
                file=file_name,
                sheet=left.title,
            )
    moved = sum(a != b for a, b in zip(left_headers, right_headers))
    ratio = moved / len(left_headers)
    report.checks[f"{file_name}/{left.title}/column_move_ratio"] = round(ratio, 6)
    if ratio < 0.6:
        report.add("column_shuffle_too_weak", "fewer than 60% of columns changed position", actual=ratio)
    report.changed = report.changed or moved > 0


def _active_header_width(sheet: Any) -> int:
    active = 0
    for col in range(1, sheet.max_column + 1):
        value = sheet.cell(1, col).value
        if value is not None and str(value).strip():
            active = col
    return active


def _validate_header_synonyms(left: Any, right: Any, file_name: str, payload: Dict[str, Any], report: IntegrityReport) -> None:
    if (left.max_row, left.max_column) != (right.max_row, right.max_column):
        return
    mappings = _declared_string_mappings(payload)
    changed = 0
    actual_pairs = set()
    for row in range(1, left.max_row + 1):
        for col in range(1, left.max_column + 1):
            old = left.cell(row, col)
            new = right.cell(row, col)
            if old.value != new.value:
                changed += 1
                actual_pairs.add((str(old.value), str(new.value)))
                if mappings.get(str(old.value)) != str(new.value):
                    report.add(
                        "undeclared_header_change",
                        "changed text is not declared in the header mapping",
                        file=file_name,
                        sheet=left.title,
                        coordinate=old.coordinate,
                        expected=mappings.get(str(old.value)),
                        actual=new.value,
                    )
                if _header_nonvalue_token(old) != _header_nonvalue_token(new):
                    report.add(
                        "header_style_changed",
                        "renamed header lost type/style/number format",
                        file=file_name,
                        sheet=left.title,
                        coordinate=old.coordinate,
                    )
            elif _data_token(old) != _data_token(new):
                report.add(
                    "non_header_cell_changed",
                    "a non-renamed cell changed value/type/formula/number format",
                    file=file_name,
                    sheet=left.title,
                    coordinate=old.coordinate,
                    expected=_data_token(old),
                    actual=_data_token(new),
                )
                if len(report.issues) >= 12:
                    break
        if len(report.issues) >= 12:
            break
    if changed == 0:
        report.add("no_header_renamed", "no declared header value changed", file=file_name, sheet=left.title)
    declared_pairs = set(_direct_mapping(payload).items())
    if actual_pairs != declared_pairs:
        report.add(
            "header_mapping_mismatch",
            "declared header mapping must exactly match actual renamed pairs",
            file=file_name,
            sheet=left.title,
            expected=sorted(declared_pairs),
            actual=sorted(actual_pairs),
        )
    report.changed = report.changed or changed > 0
    report.checks[f"{file_name}/{left.title}/renamed_cells"] = changed


def _validate_distractor_columns(left: Any, right: Any, file_name: str, payload: Dict[str, Any], report: IntegrityReport) -> None:
    if left.max_row != right.max_row:
        report.add("row_count_changed", "adding distractor columns must not change row count", file=file_name, sheet=left.title)
        return
    left_headers = [cell.value for cell in left[1]]
    right_headers = [cell.value for cell in right[1]]
    if len(set(left_headers)) != len(left_headers) or len(set(right_headers)) != len(right_headers):
        report.add("duplicate_headers", "distractor-column attack requires unique single-row headers", file=file_name, sheet=left.title)
        return
    positions = [right_headers.index(header) for header in left_headers if header in right_headers]
    if len(positions) != len(left_headers):
        missing = [header for header in left_headers if header not in right_headers]
        report.add("original_column_missing", "original columns were removed or renamed", actual=missing)
        return
    if positions != sorted(positions):
        report.add("original_columns_reordered", "original column order changed")
    added = [header for header in right_headers if header not in left_headers]
    if not added:
        report.add("distractor_column_missing", "no new distractor column was added")
    contract = _semantic_contract(payload)
    target_field = str(contract.get("target_field", "")).strip()
    distractor_field = str(contract.get("distractor_field", "")).strip()
    if added and added != [distractor_field]:
        report.add(
            "distractor_declaration_mismatch",
            "added columns must exactly match the single declared distractor field",
            file=file_name,
            sheet=left.title,
            expected=[distractor_field],
            actual=added,
        )
    if target_field in right_headers and distractor_field in right_headers:
        target_index = right_headers.index(target_field)
        distractor_index = right_headers.index(distractor_field)
        if abs(target_index - distractor_index) != 1:
            report.add(
                "distractor_not_adjacent",
                "distractor column must be immediately adjacent to its target column",
                file=file_name,
                sheet=left.title,
                expected=target_field,
                actual=distractor_field,
            )
        _validate_distractor_values(
            right,
            target_index + 1,
            distractor_index + 1,
            file_name,
            report,
        )
    for header in left_headers:
        old_col = left_headers.index(header) + 1
        new_col = right_headers.index(header) + 1
        for row in range(1, left.max_row + 1):
            old = left.cell(row, old_col)
            new = right.cell(row, new_col)
            expected = _header_token(old) if row == 1 else _data_token(old)
            actual = _header_token(new) if row == 1 else _data_token(new)
            if expected != actual:
                code = "original_header_style_lost" if row == 1 and old.value == new.value else "original_cell_changed"
                report.add(
                    code,
                    "original column value/type/formula/format changed",
                    file=file_name,
                    sheet=left.title,
                    coordinate=new.coordinate,
                    expected=expected,
                    actual=actual,
                )
                if len(report.issues) >= 12:
                    break
        if len(report.issues) >= 12:
            break
    report.changed = report.changed or bool(added)
    report.checks[f"{file_name}/{left.title}/added_columns"] = [str(item) for item in added]


def _validate_reencoding(left: Any, right: Any, file_name: str, report: IntegrityReport) -> None:
    if (left.max_row, left.max_column) != (right.max_row, right.max_column):
        return
    changed = 0
    changed_nonempty = 0
    changed_columns: Dict[int, Tuple[int, int]] = {}
    for row in range(1, left.max_row + 1):
        for col in range(1, left.max_column + 1):
            old = left.cell(row, col)
            new = right.cell(row, col)
            if old.value == new.value and _data_token(old) == _data_token(new):
                continue
            changed += 1
            total, altered = changed_columns.get(col, (0, 0))
            if old.value not in (None, ""):
                total += 1
                changed_nonempty += 1
            if old.value != new.value:
                altered += 1
            changed_columns[col] = (total, altered)
            if (old.value in (None, "")) != (new.value in (None, "")):
                report.add("missingness_changed", "NULL/empty state changed", file=file_name, sheet=left.title, coordinate=old.coordinate)
            elif not _equivalent_value(old.value, new.value):
                report.add(
                    "reencoding_not_equivalent",
                    "changed value cannot be reversibly normalized",
                    file=file_name,
                    sheet=left.title,
                    coordinate=old.coordinate,
                    expected=old.value,
                    actual=new.value,
                )
    ratios = [altered / total for total, altered in changed_columns.values() if total]
    if changed == 0:
        report.add("reencoding_missing", "no value representation changed")
    elif ratios and max(ratios) < 0.3:
        report.add("reencoding_too_weak", "no changed column reached 30% coverage", actual=max(ratios))
    report.changed = report.changed or changed > 0
    report.checks[f"{file_name}/{left.title}/changed_cells"] = changed


def _validate_unit_conversion(left: Any, right: Any, file_name: str, report: IntegrityReport) -> None:
    if (left.max_row, left.max_column) != (right.max_row, right.max_column):
        return
    ratios: List[float] = []
    label_changes = 0
    numeric_changes = 0
    for row in range(1, left.max_row + 1):
        for col in range(1, left.max_column + 1):
            old = left.cell(row, col)
            new = right.cell(row, col)
            if _data_token(old) == _data_token(new):
                continue
            if _is_number(old.value) and _is_number(new.value):
                numeric_changes += 1
                if float(old.value) != 0:
                    ratios.append(float(new.value) / float(old.value))
            elif row <= 8 and isinstance(old.value, str) and isinstance(new.value, str):
                label_changes += 1
            else:
                report.add("undeclared_unit_change", "non-numeric data changed during unit conversion", file=file_name, sheet=left.title, coordinate=old.coordinate)
    nonzero = [ratio for ratio in ratios if math.isfinite(ratio)]
    if numeric_changes == 0 or label_changes == 0:
        report.add("unit_conversion_incomplete", "numeric values and a nearby unit label must both change")
    if nonzero and not all(math.isclose(ratio, nonzero[0], rel_tol=1e-9, abs_tol=1e-12) for ratio in nonzero):
        report.add("unit_factor_inconsistent", "converted values do not share one exact factor")
    report.changed = report.changed or numeric_changes > 0
    report.checks[f"{file_name}/{left.title}/unit_factor"] = nonzero[0] if nonzero else None


def _validate_l4_semantic_contract(
    attack_type: str,
    payload: Dict[str, Any],
    report: IntegrityReport,
) -> None:
    contract = _semantic_contract(payload)
    if not contract:
        report.add("semantic_contract_missing", "L4 attacks require transformation_record.semantic_contract")
        return
    if attack_type == L4_FEATURE_PACK_ATTACK:
        pack = contract.get("feature_pack")
        if not isinstance(pack, dict):
            report.add("feature_pack_missing", "decoy feature pack must declare semantic_contract.feature_pack")
            return
        required = ("file", "misuse_answer")
        missing = [name for name in required if not _has_text(pack.get(name))]
        features = pack.get("added_features")
        target_fields = pack.get("target_fields")
        if not isinstance(pack.get("header_row"), int):
            missing.append("header_row(integer)")
        if not isinstance(target_fields, list) or not target_fields or not all(_has_text(item) for item in target_fields):
            missing.append("target_fields(nonempty list)")
        if not isinstance(features, list) or not 2 <= len(features) <= 5:
            missing.append("added_features(2..5)")
        if missing:
            report.add("feature_pack_contract_incomplete", "feature-pack declaration is incomplete", actual=missing)
            return
        names = []
        for item in features:
            if not isinstance(item, dict):
                report.add("feature_declaration_invalid", "each added feature must be an object")
                continue
            required_feature = ("name", "source_field", "noise_subtype", "exclusion_reason")
            absent = [name for name in required_feature if not _has_text(item.get(name))]
            if absent:
                report.add("feature_declaration_incomplete", "feature declaration is incomplete", actual=absent)
            names.append(str(item.get("name", "")))
        if len(set(names)) != len(names) or "" in names:
            report.add("feature_names_invalid", "added feature names must be non-empty and unique", actual=names)
        return
    if attack_type == L4_NON_OBSERVATION_ATTACK:
        declarations = contract.get("non_observation_rows")
        if not isinstance(declarations, list) or not declarations:
            report.add("non_observation_rows_missing", "non-observation injection must declare every appended row or row group")
            return
        allowed_subtypes = {"summary_row", "sample_row", "simulated_row", "check_row", "control_row"}
        for item in declarations:
            if not isinstance(item, dict):
                report.add("non_observation_declaration_invalid", "each non-observation declaration must be an object")
                continue
            required = ("file", "marker_field", "marker_value", "noise_subtype", "exclusion_reason", "misuse_answer")
            missing = [name for name in required if not _has_text(item.get(name))]
            if not _declared_rows_for_item(item):
                missing.append("row or row_start/row_end")
            if item.get("noise_subtype") not in allowed_subtypes:
                missing.append("noise_subtype")
            if missing:
                report.add("non_observation_declaration_incomplete", "non-observation declaration is incomplete", actual=missing)
        marker = contract.get("marker_column")
        if marker is not None:
            if not isinstance(marker, dict) or not _has_text(marker.get("name")) or not _has_text(marker.get("observation_value")):
                report.add("marker_column_invalid", "an appended marker_column requires name and observation_value")
        return
    if attack_type == "irrelevant_container_injection":
        declarations = contract.get("container_injections")
        if not isinstance(declarations, list) or not 1 <= len(declarations) <= 3:
            report.add("container_contract_invalid", "container_injections must declare one to three containers")
            return
        forbidden = re.compile(r"本题|答案|请勿使用|当前分析|ignore|do not use", re.IGNORECASE)
        for item in declarations:
            if not isinstance(item, dict):
                report.add("container_declaration_invalid", "each container declaration must be an object")
                continue
            required = ("kind", "file", "title", "exclusion_dimension", "exclusion_value", "misuse_result")
            missing = [name for name in required if not _has_text(item.get(name))]
            if not _has_text(item.get("source_binding_evidence")):
                missing.append("source_binding_evidence")
            if item.get("kind") not in {"file", "sheet"}:
                missing.append("kind=file|sheet")
            if item.get("kind") == "sheet" and not _has_text(item.get("sheet")):
                missing.append("sheet")
            if missing:
                report.add("container_declaration_incomplete", "container declaration is incomplete", actual=missing)
            text = " ".join(str(item.get(name, "")) for name in ("title", "exclusion_dimension", "exclusion_value"))
            if forbidden.search(text):
                report.add("task_relevant_container_metadata", "container metadata contains task relevant instructions", actual=text)
        kinds = {item.get("kind") for item in declarations if isinstance(item, dict)}
        if len(kinds) != 1:
            report.add("mixed_container_forms", "one attack may add files or sheets, not both", actual=sorted(str(item) for item in kinds))
        return

    predicates = contract.get("question_predicates")
    if not isinstance(predicates, list):
        report.add("question_predicates_missing", "row attacks must declare question_predicates")
        predicates = []
    predicate_ids = []
    for predicate in predicates:
        if not isinstance(predicate, dict):
            report.add("question_predicate_invalid", "each question predicate must be an object")
            continue
        predicate_id = str(predicate.get("id", "")).strip()
        if not predicate_id or not _has_text(predicate.get("description")) or not _has_text(predicate.get("field")):
            report.add("question_predicate_incomplete", "predicate requires id, description and field")
        predicate_ids.append(predicate_id)
    if len(set(predicate_ids)) != len(predicate_ids) or "" in predicate_ids:
        report.add("question_predicate_ids_invalid", "predicate ids must be non-empty and unique", actual=predicate_ids)
    if attack_type in {"irrelevant_record_injection", "near_match_record_injection"} and len(predicate_ids) < 2:
        report.add("insufficient_independent_predicates", "record-noise attacks require at least two independent question predicates", actual=predicate_ids)

    if attack_type in {"irrelevant_record_injection", "near_match_record_injection"}:
        records = _injected_record_declarations(contract)
        if not isinstance(records, list) or not records:
            report.add("injected_record_contract_missing", "injected_records or injected_record_groups must cover every appended row")
            return
        globally_failed = set()
        for item in records:
            if not isinstance(item, dict):
                report.add("injected_record_declaration_invalid", "each injected record must be an object")
                continue
            if not _has_text(item.get("file")) or not _declared_rows_for_item(item):
                report.add("injected_record_location_missing", "each injected declaration requires file and a row or inclusive row_start/row_end")
            results = item.get("condition_results")
            failed = item.get("failed_predicates")
            if not isinstance(results, dict) or set(results) != set(predicate_ids) or any(type(value) is not bool for value in results.values()):
                report.add("condition_results_invalid", "condition_results must contain one boolean per declared predicate", expected=predicate_ids, actual=results)
                continue
            actual_failed = {name for name, passed in results.items() if not passed}
            if not isinstance(failed, list) or set(map(str, failed)) != actual_failed:
                report.add("failed_predicates_mismatch", "failed_predicates must exactly match false condition results", expected=sorted(actual_failed), actual=failed)
            globally_failed.update(actual_failed)
            if attack_type == "irrelevant_record_injection" and len(actual_failed) < 2:
                report.add("irrelevant_record_too_close", "each irrelevant record must fail at least two independent predicates", actual=sorted(actual_failed))
            if attack_type == "near_match_record_injection" and len(actual_failed) != 1:
                report.add("near_match_failure_count_invalid", "each near-match record must fail exactly one predicate", actual=sorted(actual_failed))
        if attack_type == "near_match_record_injection" and len(globally_failed) < 2:
            report.add("near_match_failure_diversity_invalid", "near-match records must cover at least two distinct predicate ids", actual=sorted(globally_failed))
    else:
        specs = contract.get("aggregate_specs")
        if not isinstance(specs, list) or not specs:
            report.add("aggregate_specs_missing", "aggregate injection must declare every summary row in aggregate_specs")
            return
        for item in specs:
            if not isinstance(item, dict):
                report.add("aggregate_spec_invalid", "each aggregate spec must be an object")
                continue
            required_text = ("file", "marker_field", "marker_value", "aggregate_field", "operation")
            missing = [name for name in required_text if not _has_text(item.get(name))]
            if not isinstance(item.get("row"), int):
                missing.append("row")
            if item.get("operation") not in {"sum", "mean", "count", "min", "max"}:
                missing.append("operation=sum|mean|count|min|max")
            if not isinstance(item.get("filter_results"), dict) or any(value is not True for value in item.get("filter_results", {}).values()):
                missing.append("filter_results(all true)")
            if not _declared_source_rows(item):
                missing.append("source_rows/source_row_ranges")
            if missing:
                report.add("aggregate_spec_incomplete", "aggregate spec is incomplete", actual=missing)


def _validate_declared_injected_rows(
    payload: Dict[str, Any],
    file_name: str,
    sheet_name: Optional[str],
    original_last_row: int,
    attacked_last_row: int,
    report: IntegrityReport,
) -> None:
    contract = _semantic_contract(payload)
    key = "aggregate_specs" if report.attack_type == "aggregate_row_injection" else "injected_records"
    declarations = _injected_record_declarations(contract) if key == "injected_records" else contract.get(key) or []
    matching = []
    for item in declarations:
        if not isinstance(item, dict) or item.get("file") != file_name:
            continue
        declared_sheet = item.get("sheet")
        if sheet_name is None or declared_sheet in {None, "", sheet_name}:
            matching.append(item)
    expected_rows = list(range(original_last_row + 1, attacked_last_row + 1))
    actual_rows = sorted(row for item in matching for row in _declared_rows_for_item(item))
    if actual_rows != expected_rows:
        report.add(
            "injected_row_declaration_mismatch",
            "structured declarations must enumerate every physically appended row exactly once",
            file=file_name,
            sheet=sheet_name,
            expected=expected_rows,
            actual=actual_rows,
        )


def _injected_record_declarations(contract: Dict[str, Any]) -> List[Dict[str, Any]]:
    result = []
    for key in ("injected_records", "injected_record_groups"):
        values = contract.get(key)
        if isinstance(values, list):
            result.extend(item for item in values if isinstance(item, dict))
    return result


def _declared_rows_for_item(item: Dict[str, Any]) -> List[int]:
    if isinstance(item.get("row"), int):
        return [item["row"]]
    start = item.get("row_start", item.get("start_row"))
    end = item.get("row_end", item.get("end_row"))
    if isinstance(start, int) and isinstance(end, int) and start <= end:
        return list(range(start, end + 1))
    return []


def _validate_excel_marker_column(left: Any, right: Any, file_name: str, marker: Dict[str, Any], report: IntegrityReport) -> None:
    header_row = int(marker.get("header_row", 1)) if str(marker.get("header_row", 1)).isdigit() else 1
    name = str(marker.get("name", "")).strip()
    detail_value = marker.get("detail_value")
    col = right.max_column
    if not name or right.cell(header_row, col).value != name:
        report.add("marker_column_header_invalid", "appended marker column does not match its declaration", file=file_name, sheet=left.title, coordinate=right.cell(header_row, col).coordinate, expected=name, actual=right.cell(header_row, col).value)
    source_styles = {tuple(cell._style) if cell.has_style else () for cell in left[header_row]}
    marker_style = tuple(right.cell(header_row, col)._style) if right.cell(header_row, col).has_style else ()
    if marker_style not in source_styles:
        report.add("marker_header_style_invalid", "marker header must use an existing header style", file=file_name, sheet=left.title, coordinate=right.cell(header_row, col).coordinate)
    for row in range(header_row + 1, left.max_row + 1):
        if right.cell(row, col).value != detail_value:
            report.add("detail_marker_invalid", "every original data row must receive the declared detail marker", file=file_name, sheet=left.title, coordinate=right.cell(row, col).coordinate, expected=detail_value, actual=right.cell(row, col).value)
            if len(report.issues) >= 12:
                break


def _validate_csv_marker_column(left: List[List[str]], right: List[List[str]], file_name: str, marker: Dict[str, Any], report: IntegrityReport) -> None:
    name = str(marker.get("name", ""))
    detail = str(marker.get("detail_value", ""))
    if right[0][-1] != name:
        report.add("marker_column_header_invalid", "CSV marker column does not match declaration", file=file_name, expected=name, actual=right[0][-1])
    for row_number, row in enumerate(right[1 : len(left)], 2):
        if len(row) != len(right[0]) or row[-1] != detail:
            report.add("detail_marker_invalid", "each original CSV row must receive the detail marker", file=file_name, coordinate=str(row_number), expected=detail, actual=row[-1] if row else None)
            if len(report.issues) >= 12:
                break


def _declared_source_rows(spec: Dict[str, Any]) -> List[int]:
    rows = []
    for value in spec.get("source_rows") or []:
        if isinstance(value, int):
            rows.append(value)
    for bounds in spec.get("source_row_ranges") or []:
        if isinstance(bounds, list) and len(bounds) == 2 and all(isinstance(value, int) for value in bounds) and bounds[0] <= bounds[1]:
            rows.extend(range(bounds[0], bounds[1] + 1))
    return sorted(set(rows))


def _computed_aggregate(values: List[Any], operation: str) -> Optional[float]:
    if operation == "count":
        return float(sum(value is not None and str(value).strip() != "" for value in values))
    numbers = [_parse_equivalent_number(value) for value in values]
    numbers = [value for value in numbers if value is not None]
    if not numbers:
        return None
    if operation == "sum":
        return sum(numbers)
    if operation == "mean":
        return sum(numbers) / len(numbers)
    if operation == "min":
        return min(numbers)
    if operation == "max":
        return max(numbers)
    return None


def _validate_excel_aggregate_specs(left: Any, right: Any, file_name: str, payload: Dict[str, Any], report: IntegrityReport) -> None:
    specs = _semantic_contract(payload).get("aggregate_specs") or []
    for spec in specs:
        if not isinstance(spec, dict) or spec.get("file") != file_name or spec.get("sheet") not in {None, "", left.title}:
            continue
        header_row = int(spec.get("header_row", 1)) if str(spec.get("header_row", 1)).isdigit() else 1
        headers = {str(left.cell(header_row, col).value): col for col in range(1, left.max_column + 1)}
        aggregate_col = headers.get(str(spec.get("aggregate_field")))
        marker_headers = {str(right.cell(header_row, col).value): col for col in range(1, right.max_column + 1)}
        marker_col = marker_headers.get(str(spec.get("marker_field")))
        row = spec.get("row")
        if aggregate_col is None or marker_col is None or not isinstance(row, int) or row > right.max_row:
            report.add("aggregate_spec_location_invalid", "aggregate or marker field/row cannot be located", file=file_name, sheet=left.title, actual=spec)
            continue
        if right.cell(row, marker_col).value != spec.get("marker_value"):
            report.add("summary_marker_invalid", "summary row marker does not match declaration", file=file_name, sheet=left.title, coordinate=right.cell(row, marker_col).coordinate, expected=spec.get("marker_value"), actual=right.cell(row, marker_col).value)
        source_rows = _declared_source_rows(spec)
        if any(source_row <= header_row or source_row > left.max_row for source_row in source_rows):
            report.add("aggregate_source_row_invalid", "aggregate source rows must reference original data rows", file=file_name, sheet=left.title)
            continue
        expected = _computed_aggregate([left.cell(source_row, aggregate_col).value for source_row in source_rows], str(spec.get("operation")))
        actual = _parse_equivalent_number(right.cell(row, aggregate_col).value)
        if expected is None or actual is None or not math.isclose(expected, actual, rel_tol=1e-9, abs_tol=1e-8):
            report.add("aggregate_value_incorrect", "declared summary value does not recompute from original source rows", file=file_name, sheet=left.title, coordinate=right.cell(row, aggregate_col).coordinate, expected=expected, actual=actual)
        report.checks[f"{file_name}/{left.title}/{row}/aggregate_recomputed"] = expected


def _validate_csv_aggregate_specs(left: List[List[str]], right: List[List[str]], file_name: str, payload: Dict[str, Any], report: IntegrityReport) -> None:
    specs = _semantic_contract(payload).get("aggregate_specs") or []
    headers = {name: index for index, name in enumerate(right[0])} if right else {}
    for spec in specs:
        if not isinstance(spec, dict) or spec.get("file") != file_name:
            continue
        aggregate_col = headers.get(str(spec.get("aggregate_field")))
        marker_col = headers.get(str(spec.get("marker_field")))
        row_number = spec.get("row")
        if aggregate_col is None or marker_col is None or not isinstance(row_number, int) or row_number > len(right):
            report.add("aggregate_spec_location_invalid", "CSV aggregate or marker field/row cannot be located", file=file_name, actual=spec)
            continue
        row = right[row_number - 1]
        if row[marker_col] != str(spec.get("marker_value")):
            report.add("summary_marker_invalid", "CSV summary marker does not match declaration", file=file_name, coordinate=str(row_number))
        source_rows = _declared_source_rows(spec)
        if any(source_row < 2 or source_row > len(left) for source_row in source_rows):
            report.add("aggregate_source_row_invalid", "CSV aggregate source rows must reference original data rows", file=file_name)
            continue
        expected = _computed_aggregate([left[source_row - 1][aggregate_col] for source_row in source_rows], str(spec.get("operation")))
        actual = _parse_equivalent_number(row[aggregate_col])
        if expected is None or actual is None or not math.isclose(expected, actual, rel_tol=1e-9, abs_tol=1e-8):
            report.add("aggregate_value_incorrect", "CSV summary value does not recompute from original source rows", file=file_name, coordinate=str(row_number), expected=expected, actual=actual)


def _validate_declared_containers(
    payload: Dict[str, Any],
    *,
    added_files: List[str],
    added_sheets: List[str],
    file_name: Optional[str],
    report: IntegrityReport,
) -> None:
    declarations = _semantic_contract(payload).get("container_injections") or []
    if file_name is None:
        declared = sorted(str(item.get("file")) for item in declarations if isinstance(item, dict) and item.get("kind") == "file")
        if declared != sorted(added_files):
            report.add("container_declaration_mismatch", "declared added files do not match the output package", expected=declared, actual=sorted(added_files))
    else:
        declared = sorted(
            str(item.get("sheet"))
            for item in declarations
            if isinstance(item, dict) and item.get("kind") == "sheet" and item.get("file") == file_name
        )
        if declared != sorted(added_sheets):
            report.add("container_declaration_mismatch", "declared added sheets do not match the workbook", file=file_name, expected=declared, actual=sorted(added_sheets))


def _feature_pack_for(payload: Dict[str, Any], file_name: str, sheet_name: Optional[str]) -> Optional[Dict[str, Any]]:
    pack = _semantic_contract(payload).get("feature_pack")
    if not isinstance(pack, dict) or pack.get("file") != file_name:
        return None
    declared_sheet = pack.get("sheet")
    if sheet_name is not None and declared_sheet not in {sheet_name}:
        return None
    return pack


def _non_observation_for(
    payload: Dict[str, Any], file_name: str, sheet_name: Optional[str]
) -> List[Dict[str, Any]]:
    declarations = _semantic_contract(payload).get("non_observation_rows") or []
    result = []
    for item in declarations:
        if not isinstance(item, dict) or item.get("file") != file_name:
            continue
        declared_sheet = item.get("sheet")
        if sheet_name is None:
            if declared_sheet not in {None, ""}:
                continue
        elif declared_sheet != sheet_name:
            continue
        result.append(item)
    return result


def _validate_excel_decoy_feature_pack(
    left: Any, right: Any, file_name: str, payload: Dict[str, Any], report: IntegrityReport
) -> None:
    pack = _feature_pack_for(payload, file_name, left.title)
    if pack is None:
        if _sheet_matrix_token(left) != _sheet_matrix_token(right):
            report.add("undeclared_sheet_changed", "feature pack changed a sheet outside its declaration", file=file_name, sheet=left.title)
        return
    header_row = pack.get("header_row")
    if header_row != 1:
        report.add("feature_pack_header_row_invalid", "feature pack currently requires header_row=1", file=file_name, sheet=left.title, actual=header_row)
        return
    features = pack.get("added_features") or []
    if right.max_row != left.max_row or right.max_column != left.max_column + len(features):
        report.add(
            "feature_pack_shape_invalid",
            "feature pack may only append its declared columns without changing rows",
            file=file_name,
            sheet=left.title,
            expected=(left.max_row, left.max_column + len(features)),
            actual=(right.max_row, right.max_column),
        )
        return
    _validate_original_cells_preserved(left, right, file_name, report)
    old_headers = [left.cell(1, col).value for col in range(1, left.max_column + 1)]
    new_headers = [right.cell(1, col).value for col in range(left.max_column + 1, right.max_column + 1)]
    declared_names = [str(item.get("name", "")) for item in features if isinstance(item, dict)]
    if new_headers != declared_names:
        report.add("feature_pack_declaration_mismatch", "appended headers must exactly match declared feature names", file=file_name, sheet=left.title, expected=declared_names, actual=new_headers)
    if len(set(old_headers + new_headers)) != len(old_headers) + len(new_headers):
        report.add("duplicate_headers", "feature pack requires unique headers", file=file_name, sheet=left.title)
    header_styles = {tuple(cell._style) if cell.has_style else () for cell in left[1]}
    for offset, feature in enumerate(features, start=left.max_column + 1):
        if not isinstance(feature, dict):
            continue
        source_field = str(feature.get("source_field", ""))
        if source_field not in old_headers:
            report.add("feature_source_missing", "declared source field is absent from the original sheet", file=file_name, sheet=left.title, actual=source_field)
            continue
        style = tuple(right.cell(1, offset)._style) if right.cell(1, offset).has_style else ()
        if style not in header_styles:
            report.add("feature_header_style_invalid", "appended feature header must reuse an original header style", file=file_name, sheet=left.title, coordinate=right.cell(1, offset).coordinate)
        source_col = old_headers.index(source_field) + 1
        source_families = {_excel_value_family(left.cell(row, source_col).value) for row in range(2, left.max_row + 1)} - {"null"}
        decoy_values = [right.cell(row, offset).value for row in range(2, right.max_row + 1)]
        decoy_families = {_excel_value_family(value) for value in decoy_values} - {"null"}
        coverage = sum(value is not None for value in decoy_values) / max(1, len(decoy_values))
        if coverage < 0.6 or not decoy_families:
            report.add("feature_pack_sparse", "each decoy feature must cover at least 60% of data rows", file=file_name, sheet=left.title, coordinate=right.cell(1, offset).coordinate, actual=coverage)
        if source_families and not decoy_families.issubset(source_families):
            report.add("feature_type_incompatible", "decoy feature values are not type-compatible with their declared source field", file=file_name, sheet=left.title, coordinate=right.cell(1, offset).coordinate, expected=sorted(source_families), actual=sorted(decoy_families))
    report.changed = True
    report.checks[f"{file_name}/{left.title}/added_features"] = new_headers


def _validate_excel_non_observation_rows(
    left: Any, right: Any, file_name: str, payload: Dict[str, Any], report: IntegrityReport
) -> None:
    declarations = _non_observation_for(payload, file_name, left.title)
    if not declarations:
        if _sheet_matrix_token(left) != _sheet_matrix_token(right):
            report.add("undeclared_sheet_changed", "non-observation injection changed a sheet outside its declaration", file=file_name, sheet=left.title)
        return
    _validate_excel_non_observation_table(left, right, file_name, payload, declarations, report)


def _validate_excel_non_observation_table(
    left: Any,
    right: Any,
    file_name: str,
    payload: Dict[str, Any],
    declarations: List[Dict[str, Any]],
    report: IntegrityReport,
) -> None:
    marker = _semantic_contract(payload).get("marker_column")
    marker = marker if isinstance(marker, dict) else None
    expected_columns = left.max_column + (1 if marker else 0)
    if right.max_column != expected_columns or right.max_row <= left.max_row:
        report.add("non_observation_shape_invalid", "non-observation rows must be appended; only one declared marker column may be appended", file=file_name, sheet=left.title, expected=f"rows>{left.max_row}, columns={expected_columns}", actual=f"rows={right.max_row}, columns={right.max_column}")
        return
    _validate_original_cells_preserved(left, right, file_name, report)
    if marker:
        _validate_excel_observation_marker_column(left, right, file_name, marker, report)
    _validate_declared_non_observation_rows(declarations, file_name, left.title, left.max_row, right.max_row, report)
    headers = {str(right.cell(1, col).value): col for col in range(1, right.max_column + 1)}
    _validate_non_observation_marker_values(right, headers, declarations, file_name, left.title, report)
    _validate_excel_appended_row_types(left, right, file_name, report)
    _validate_non_observation_count(left.max_row - 1, right.max_row - left.max_row, file_name, left.title, report)
    report.changed = True
    report.checks[f"{file_name}/{left.title}/added_non_observation_rows"] = right.max_row - left.max_row


def _validate_original_cells_preserved(left: Any, right: Any, file_name: str, report: IntegrityReport) -> None:
    for row in range(1, left.max_row + 1):
        for col in range(1, left.max_column + 1):
            old = left.cell(row, col)
            new = right.cell(row, col)
            token = _header_token if row == 1 else _data_token
            if token(old) != token(new):
                report.add("original_records_changed", "original cells must remain at the same coordinates", file=file_name, sheet=left.title, coordinate=new.coordinate, expected=token(old), actual=token(new))
                if len(report.issues) >= 12:
                    return


def _validate_excel_observation_marker_column(left: Any, right: Any, file_name: str, marker: Dict[str, Any], report: IntegrityReport) -> None:
    name = str(marker.get("name", "")).strip()
    observation_value = marker.get("observation_value")
    column = right.max_column
    if right.cell(1, column).value != name:
        report.add("marker_column_header_invalid", "appended marker column does not match its declaration", file=file_name, sheet=left.title, coordinate=right.cell(1, column).coordinate, expected=name, actual=right.cell(1, column).value)
    styles = {tuple(cell._style) if cell.has_style else () for cell in left[1]}
    actual_style = tuple(right.cell(1, column)._style) if right.cell(1, column).has_style else ()
    if actual_style not in styles:
        report.add("marker_header_style_invalid", "marker header must reuse an original header style", file=file_name, sheet=left.title, coordinate=right.cell(1, column).coordinate)
    for row in range(2, left.max_row + 1):
        if right.cell(row, column).value != observation_value:
            report.add("observation_marker_invalid", "every original data row must receive the declared observation marker", file=file_name, sheet=left.title, coordinate=right.cell(row, column).coordinate, expected=observation_value, actual=right.cell(row, column).value)
            if len(report.issues) >= 12:
                return


def _validate_excel_appended_row_types(left: Any, right: Any, file_name: str, report: IntegrityReport) -> None:
    for col in range(1, left.max_column + 1):
        expected = {_excel_value_family(left.cell(row, col).value) for row in range(2, left.max_row + 1)} - {"null"}
        actual = {_excel_value_family(right.cell(row, col).value) for row in range(left.max_row + 1, right.max_row + 1)} - {"null"}
        if expected and actual and not actual.issubset(expected):
            report.add("injected_row_type_incompatible", "appended record values are not schema-compatible with the original column", file=file_name, sheet=left.title, coordinate=right.cell(1, col).coordinate, expected=sorted(expected), actual=sorted(actual))


def _validate_declared_non_observation_rows(
    declarations: List[Dict[str, Any]], file_name: str, sheet_name: Optional[str], original_last_row: int, attacked_last_row: int, report: IntegrityReport
) -> None:
    expected = list(range(original_last_row + 1, attacked_last_row + 1))
    actual = sorted(row for item in declarations for row in _declared_rows_for_item(item))
    if actual != expected:
        report.add("non_observation_row_declaration_mismatch", "declarations must enumerate every appended physical row exactly once", file=file_name, sheet=sheet_name, expected=expected, actual=actual)


def _validate_non_observation_marker_values(
    sheet: Any, headers: Dict[str, int], declarations: List[Dict[str, Any]], file_name: str, sheet_name: Optional[str], report: IntegrityReport
) -> None:
    for item in declarations:
        marker_field = str(item.get("marker_field", ""))
        marker_col = headers.get(marker_field)
        if marker_col is None:
            report.add("non_observation_marker_missing", "declared marker field cannot be located", file=file_name, sheet=sheet_name, actual=marker_field)
            continue
        for row in _declared_rows_for_item(item):
            if sheet.cell(row, marker_col).value != item.get("marker_value"):
                report.add("non_observation_marker_invalid", "appended row marker does not match its declaration", file=file_name, sheet=sheet_name, coordinate=sheet.cell(row, marker_col).coordinate, expected=item.get("marker_value"), actual=sheet.cell(row, marker_col).value)


def _validate_non_observation_count(original_rows: int, added_rows: int, file_name: str, sheet_name: Optional[str], report: IntegrityReport) -> None:
    maximum = max(5, math.ceil(max(1, original_rows) * 0.10))
    if not 1 <= added_rows <= maximum:
        report.add("non_observation_count_invalid", "inject between one row and 10% of observations (up to five rows for small tables)", file=file_name, sheet=sheet_name, expected=f"1..{maximum}", actual=added_rows)


def _excel_value_family(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return "number"
    if isinstance(value, (date, datetime)):
        return "date"
    return "text"


def _validate_row_injection(
    left: Any,
    right: Any,
    file_name: str,
    attack_type: str,
    payload: Dict[str, Any],
    report: IntegrityReport,
) -> None:
    if _sheet_matrix_token(left) == _sheet_matrix_token(right):
        return
    marker = _semantic_contract(payload).get("marker_column") if attack_type == "aggregate_row_injection" else None
    marker = marker if isinstance(marker, dict) else None
    expected_columns = left.max_column + (1 if marker else 0)
    if right.max_column != expected_columns or right.max_row <= left.max_row:
        report.add(
            "row_injection_shape_invalid",
            "rows must be appended; only aggregate injection may append one declared marker column",
            file=file_name,
            sheet=left.title,
            expected=f"rows>{left.max_row}, columns={expected_columns}",
            actual=f"rows={right.max_row}, columns={right.max_column}",
        )
        return
    for row in range(1, left.max_row + 1):
        for col in range(1, left.max_column + 1):
            old = left.cell(row, col)
            new = right.cell(row, col)
            token = _header_token if row == 1 else _data_token
            if token(old) != token(new):
                report.add(
                    "original_records_changed",
                    "original cells must remain at the same coordinates",
                    file=file_name,
                    sheet=left.title,
                    coordinate=new.coordinate,
                    expected=token(old),
                    actual=token(new),
                )
                if len(report.issues) >= 12:
                    break
        if len(report.issues) >= 12:
            break
    if marker:
        _validate_excel_marker_column(left, right, file_name, marker, report)
    added = right.max_row - left.max_row
    original = max(1, left.max_row - 1)
    ratio = added / original
    if attack_type == "irrelevant_record_injection" and not 0.1 <= ratio <= 0.3:
        report.add("injection_count_invalid", "irrelevant record injection must add 10%-30%", actual=ratio)
    if attack_type == "near_match_record_injection" and ratio > 0.3:
        report.add("injection_count_invalid", "near-match injection may not exceed 30% of the full table", actual=ratio)
    _validate_declared_injected_rows(payload, file_name, left.title, left.max_row, right.max_row, report)
    if attack_type == "aggregate_row_injection":
        _validate_excel_aggregate_specs(left, right, file_name, payload, report)
    report.changed = report.changed or added > 0
    report.checks[f"{file_name}/{left.title}/added_rows"] = added


def _validate_hierarchical_relayout(left: Any, right: Any, file_name: str, payload: Dict[str, Any], report: IntegrityReport) -> None:
    left_formulas = Counter(_formula_token(cell) for row in left.iter_rows() for cell in row if cell.data_type == "f")
    right_formulas = Counter(_formula_token(cell) for row in right.iter_rows() for cell in row if cell.data_type == "f")
    if left_formulas != right_formulas:
        report.add("formula_set_changed", "formula cells changed during header relayout", file=file_name, sheet=left.title)
    spec = _hierarchical_sheet_spec(payload, file_name, left.title)
    if not spec:
        report.add("header_relayout_spec_missing", "hierarchical relayout must declare a sheet_specs entry", file=file_name, sheet=left.title)
    else:
        try:
            original_start = int(spec["original_data_start_row"])
            final_start = int(spec["final_data_start_row"])
        except (KeyError, TypeError, ValueError):
            report.add("data_start_row_invalid", "sheet spec requires integer original/final data start rows", file=file_name, sheet=left.title)
        else:
            if left.max_column != right.max_column:
                report.add("data_column_count_changed", "header relayout must preserve data column count", file=file_name, sheet=left.title, expected=left.max_column, actual=right.max_column)
            left_count = max(0, left.max_row - original_start + 1)
            right_count = max(0, right.max_row - final_start + 1)
            if left_count != right_count:
                report.add("data_row_count_changed", "header relayout changed data row count", file=file_name, sheet=left.title, expected=left_count, actual=right_count)
            for offset in range(min(left_count, right_count)):
                for col in range(1, min(left.max_column, right.max_column) + 1):
                    old = left.cell(original_start + offset, col)
                    new = right.cell(final_start + offset, col)
                    if _data_token(old) != _data_token(new):
                        report.add(
                            "data_cell_changed",
                            "data value/type/formula/number format changed during header relayout",
                            file=file_name,
                            sheet=left.title,
                            coordinate=new.coordinate,
                            expected=_data_token(old),
                            actual=_data_token(new),
                        )
                        if len(report.issues) >= 12:
                            break
                if len(report.issues) >= 12:
                    break
    if sorted(str(item) for item in left.merged_cells.ranges) == sorted(str(item) for item in right.merged_cells.ranges):
        report.add("header_hierarchy_not_changed", "merged/header hierarchy did not change", file=file_name, sheet=left.title)
    report.changed = True


def _validate_excel_containers(
    original_wb: Any,
    attacked_wb: Any,
    attack_type: str,
    file_name: str,
    payload: Dict[str, Any],
    report: IntegrityReport,
) -> None:
    if attack_type == "irrelevant_container_injection":
        missing = [name for name in original_wb.sheetnames if name not in attacked_wb.sheetnames]
        if missing:
            report.add("original_sheet_missing", "original sheets were removed", file=file_name, actual=missing)
        for name in original_wb.sheetnames:
            if name not in attacked_wb.sheetnames:
                continue
            if _sheet_matrix_token(original_wb[name]) != _sheet_matrix_token(attacked_wb[name]):
                report.add("original_container_changed", "original sheet content changed", file=file_name, sheet=name)
        added = [name for name in attacked_wb.sheetnames if name not in original_wb.sheetnames]
        declarations = _semantic_contract(payload).get("container_injections") or []
        declared_for_file = [
            item
            for item in declarations
            if isinstance(item, dict) and item.get("kind") == "sheet" and item.get("file") == file_name
        ]
        if declared_for_file and not 1 <= len(added) <= 3:
            report.add("container_count_invalid", "one to three declared sheets must be added", actual=len(added))
        _validate_declared_containers(payload, added_files=[], added_sheets=added, file_name=file_name, report=report)
        by_sheet = {
            str(item.get("sheet")): item
            for item in declarations
            if isinstance(item, dict) and item.get("kind") == "sheet" and item.get("file") == file_name
        }
        for name in added:
            declaration = by_sheet.get(name, {})
            evidence = " ".join(
                str(attacked_wb[name].cell(row, col).value or "")
                for row in range(1, min(attacked_wb[name].max_row, 8) + 1)
                for col in range(1, min(attacked_wb[name].max_column, 8) + 1)
            )
            title = str(declaration.get("title", ""))
            exclusion = str(declaration.get("exclusion_value", ""))
            if title not in name and title not in evidence:
                report.add("container_title_missing", "declared neutral title is not present in the added sheet", file=file_name, sheet=name, expected=title)
            if exclusion not in name and exclusion not in evidence:
                report.add("container_exclusion_metadata_missing", "declared exclusion metadata is not present in the added sheet", file=file_name, sheet=name, expected=exclusion)
        report.changed = bool(added)
        report.checks[f"{file_name}/added_sheets"] = added
        return
    original_rows = _canonical_workbook_rows(original_wb)
    attacked_rows = _canonical_workbook_rows(attacked_wb)
    if original_rows != attacked_rows:
        report.add("cross_sheet_records_changed", "canonical record multiset changed across sheets", file=file_name)
    if len(original_wb.sheetnames) == len(attacked_wb.sheetnames):
        report.add("sheet_count_not_changed", "cross-sheet attack must split into more sheets or merge into fewer sheets", file=file_name, expected=len(original_wb.sheetnames), actual=len(attacked_wb.sheetnames))
    for sheet in attacked_wb.worksheets:
        headers = [cell.value for cell in sheet[1]]
        if sheet.max_row < 2 or len(set(headers)) != len(headers):
            report.add("invalid_relayout_sheet", "each relocated sheet must have unique headers and at least one data row", file=file_name, sheet=sheet.title)
    original_formulas = Counter(_formula_token(cell) for sheet in original_wb.worksheets for row in sheet.iter_rows() for cell in row if cell.data_type == "f")
    attacked_formulas = Counter(_formula_token(cell) for sheet in attacked_wb.worksheets for row in sheet.iter_rows() for cell in row if cell.data_type == "f")
    if original_formulas != attacked_formulas:
        report.add("formula_set_changed", "cross-sheet relayout changed formula text or number formats", file=file_name)
    original_header_styles = _header_styles_by_name(original_wb)
    attacked_header_styles = _header_styles_by_name(attacked_wb)
    for header, styles in attacked_header_styles.items():
        if header in original_header_styles and not styles.issubset(original_header_styles[header]):
            report.add(
                "header_style_changed",
                f"header style changed while relocating field {header!r}",
                file=file_name,
            )
    report.changed = original_wb.sheetnames != attacked_wb.sheetnames


def _validate_preserved_or_changed(left: Any, right: Any, file_name: str, report: IntegrityReport) -> None:
    left_token = _sheet_matrix_token(left)
    right_token = _sheet_matrix_token(right)
    report.changed = report.changed or left_token != right_token


def _validate_delimited_pair(
    original_path: Path,
    attacked_path: Path,
    attack_type: str,
    payload: Dict[str, Any],
    selection: Dict[str, Any],
    report: IntegrityReport,
    delimiter: str,
) -> None:
    left = _read_delimited(original_path, delimiter)
    right = _read_delimited(attacked_path, delimiter)
    if attack_type == "row_order_shuffle":
        if not left or not right or left[0] != right[0] or Counter(map(tuple, left[1:])) != Counter(map(tuple, right[1:])):
            report.add("data_row_multiset_changed", "CSV row shuffle changed header or lexical row values", file=original_path.name)
            return
        moved = sum(a != b for a, b in zip(left[1:], right[1:]))
        ratio = moved / max(1, len(left) - 1)
        if ratio < 0.8:
            report.add("row_shuffle_too_weak", "fewer than 80% of CSV rows moved", actual=ratio)
        report.changed = moved > 0
    elif attack_type == "column_order_shuffle":
        _validate_csv_column_shuffle(left, right, original_path.name, report)
    elif attack_type == "header_synonym_substitution":
        _validate_csv_header_synonyms(left, right, original_path.name, payload, report)
    elif attack_type == "semantic_distractor_column":
        _validate_csv_distractors(left, right, original_path.name, payload, report)
    elif attack_type == L4_FEATURE_PACK_ATTACK:
        _validate_csv_decoy_feature_pack(left, right, original_path.name, payload, report)
    elif attack_type == L4_NON_OBSERVATION_ATTACK:
        _validate_csv_non_observation_rows(left, right, original_path.name, payload, report)
    elif attack_type == "csv_wide_long_reshape":
        _validate_csv_reshape(left, right, original_path.name, payload, report)
    elif attack_type in {"equivalent_value_reencoding", "unit_scale_conversion"}:
        report.changed = left != right
        if not report.changed:
            report.add("attack_not_changed", "CSV values did not change", file=original_path.name)
    elif left != right:
        report.changed = True


def _validate_csv_column_shuffle(left: List[List[str]], right: List[List[str]], file_name: str, report: IntegrityReport) -> None:
    if not left or not right or len(left) != len(right):
        report.add("csv_shape_changed", "CSV row count changed", file=file_name)
        return
    lh, rh = left[0], right[0]
    if len(set(lh)) != len(lh) or set(lh) != set(rh):
        report.add("header_set_changed", "CSV headers changed or are not unique", file=file_name)
        return
    for header in lh:
        if [row[lh.index(header)] for row in left] != [row[rh.index(header)] for row in right]:
            report.add("column_content_changed", f"CSV column changed for {header!r}", file=file_name)
    ratio = sum(a != b for a, b in zip(lh, rh)) / len(lh)
    if ratio < 0.6:
        report.add("column_shuffle_too_weak", "fewer than 60% of CSV columns moved", actual=ratio)
    report.changed = ratio > 0


def _validate_csv_header_synonyms(left: List[List[str]], right: List[List[str]], file_name: str, payload: Dict[str, Any], report: IntegrityReport) -> None:
    if len(left) != len(right) or not left or len(left[0]) != len(right[0]):
        report.add("csv_shape_changed", "header synonym attack changed CSV shape", file=file_name)
        return
    mappings = _declared_string_mappings(payload)
    changed = 0
    for old, new in zip(left[0], right[0]):
        if old != new:
            changed += 1
            if mappings.get(old) != new:
                report.add("undeclared_header_change", "CSV header change was not declared", expected=mappings.get(old), actual=new)
    if left[1:] != right[1:]:
        report.add("non_header_cell_changed", "CSV data changed during header substitution", file=file_name)
    report.changed = changed > 0


def _validate_csv_distractors(
    left: List[List[str]],
    right: List[List[str]],
    file_name: str,
    payload: Dict[str, Any],
    report: IntegrityReport,
) -> None:
    if len(left) != len(right) or not left or not right:
        report.add("csv_row_count_changed", "adding columns changed CSV row count", file=file_name)
        return
    lh, rh = left[0], right[0]
    if any(header not in rh for header in lh):
        report.add("original_column_missing", "original CSV column missing", file=file_name)
        return
    for header in lh:
        if [row[lh.index(header)] for row in left] != [row[rh.index(header)] for row in right]:
            report.add("original_cell_changed", f"original CSV column changed for {header!r}", file=file_name)
    added = [header for header in rh if header not in lh]
    if not added:
        report.add("distractor_column_missing", "no CSV distractor column added", file=file_name)
    contract = _semantic_contract(payload)
    target_field = str(contract.get("target_field", "")).strip()
    distractor_field = str(contract.get("distractor_field", "")).strip()
    if added and added != [distractor_field]:
        report.add("distractor_declaration_mismatch", "CSV must add exactly the declared distractor", expected=[distractor_field], actual=added)
    if target_field in rh and distractor_field in rh:
        target_index = rh.index(target_field)
        distractor_index = rh.index(distractor_field)
        if abs(target_index - distractor_index) != 1:
            report.add("distractor_not_adjacent", "CSV distractor must be adjacent to its target", file=file_name)
        target_values = [row[target_index] for row in right[1:] if len(row) > target_index]
        distractor_values = [row[distractor_index] for row in right[1:] if len(row) > distractor_index]
        _validate_lexical_distractor_values(target_values, distractor_values, file_name, report)
    report.changed = bool(added)


def _validate_csv_decoy_feature_pack(
    left: List[List[str]], right: List[List[str]], file_name: str, payload: Dict[str, Any], report: IntegrityReport
) -> None:
    pack = _feature_pack_for(payload, file_name, None)
    if pack is None:
        if left != right:
            report.add("undeclared_file_changed", "feature pack changed a file outside its declaration", file=file_name)
        return
    features = pack.get("added_features") or []
    if not left or not right or len(left) != len(right) or len(right[0]) != len(left[0]) + len(features):
        report.add("feature_pack_shape_invalid", "CSV feature pack may only append its declared columns without changing rows", file=file_name)
        return
    left_header, right_header = left[0], right[0]
    if any(row[: len(left_header)] != old for row, old in zip(right, left)):
        report.add("original_records_changed", "original CSV cells must remain unchanged", file=file_name)
    names = [str(item.get("name", "")) for item in features if isinstance(item, dict)]
    added = right_header[len(left_header) :]
    if added != names:
        report.add("feature_pack_declaration_mismatch", "appended CSV headers must exactly match declared feature names", file=file_name, expected=names, actual=added)
    if len(set(right_header)) != len(right_header):
        report.add("duplicate_headers", "feature pack requires unique CSV headers", file=file_name)
    for index, feature in enumerate(features, start=len(left_header)):
        if not isinstance(feature, dict):
            continue
        source = str(feature.get("source_field", ""))
        if source not in left_header:
            report.add("feature_source_missing", "declared source field is absent from the original CSV", file=file_name, actual=source)
            continue
        values = [row[index] for row in right[1:] if len(row) > index]
        coverage = sum(value != "" for value in values) / max(1, len(values))
        if coverage < 0.6:
            report.add("feature_pack_sparse", "each CSV decoy feature must cover at least 60% of data rows", file=file_name, coordinate=right_header[index], actual=coverage)
    report.changed = True
    report.checks[f"{file_name}/added_features"] = added


def _validate_csv_non_observation_rows(
    left: List[List[str]], right: List[List[str]], file_name: str, payload: Dict[str, Any], report: IntegrityReport
) -> None:
    declarations = _non_observation_for(payload, file_name, None)
    if not declarations:
        if left != right:
            report.add("undeclared_file_changed", "non-observation injection changed a file outside its declaration", file=file_name)
        return
    marker = _semantic_contract(payload).get("marker_column")
    marker = marker if isinstance(marker, dict) else None
    if not left or not right:
        report.add("non_observation_shape_invalid", "CSV input and output must be non-empty", file=file_name)
        return
    expected_width = len(left[0]) + (1 if marker else 0)
    if len(right[0]) != expected_width or len(right) <= len(left):
        report.add("non_observation_shape_invalid", "non-observation rows must be appended; only one declared marker column may be appended", file=file_name)
        return
    if right[0][: len(left[0])] != left[0] or any(row[: len(left[0])] != old for row, old in zip(right[1 : len(left)], left[1:])):
        report.add("original_records_changed", "original CSV records must remain at the same positions", file=file_name)
    if marker:
        marker_name = str(marker.get("name", ""))
        if right[0][-1] != marker_name:
            report.add("marker_column_header_invalid", "appended marker column does not match its declaration", file=file_name, expected=marker_name, actual=right[0][-1])
        for row_number, row in enumerate(right[1 : len(left)], start=2):
            if row[-1] != str(marker.get("observation_value")):
                report.add("observation_marker_invalid", "every original CSV row must receive the observation marker", file=file_name, coordinate=str(row_number), expected=marker.get("observation_value"), actual=row[-1])
                break
    _validate_declared_non_observation_rows(declarations, file_name, None, len(left), len(right), report)
    headers = {name: index for index, name in enumerate(right[0])}
    for item in declarations:
        marker_col = headers.get(str(item.get("marker_field", "")))
        if marker_col is None:
            report.add("non_observation_marker_missing", "declared marker field cannot be located", file=file_name, actual=item.get("marker_field"))
            continue
        for row_number in _declared_rows_for_item(item):
            if right[row_number - 1][marker_col] != str(item.get("marker_value")):
                report.add("non_observation_marker_invalid", "appended CSV row marker does not match its declaration", file=file_name, coordinate=str(row_number), expected=item.get("marker_value"), actual=right[row_number - 1][marker_col])
    _validate_non_observation_count(len(left) - 1, len(right) - len(left), file_name, None, report)
    report.changed = True


def _validate_csv_row_injection(
    left: List[List[str]],
    right: List[List[str]],
    file_name: str,
    attack_type: str,
    payload: Dict[str, Any],
    report: IntegrityReport,
) -> None:
    marker = _semantic_contract(payload).get("marker_column") if attack_type == "aggregate_row_injection" else None
    marker = marker if isinstance(marker, dict) else None
    expected_width = len(left[0]) + (1 if marker else 0) if left else 0
    if not left or not right or len(right[0]) != expected_width:
        report.add("header_changed", "CSV row injection changed header unexpectedly", file=file_name)
        return
    if right[0][: len(left[0])] != left[0] or any(row[: len(left[0])] != old for row, old in zip(right[1 : len(left)], left[1:])):
        report.add("original_records_changed", "original CSV records changed", file=file_name)
    if marker:
        _validate_csv_marker_column(left, right, file_name, marker, report)
    added = len(right) - len(left)
    ratio = added / max(1, len(left) - 1)
    if added <= 0:
        report.add("rows_not_added", "no CSV rows added", file=file_name)
    if attack_type == "irrelevant_record_injection" and not 0.1 <= ratio <= 0.3:
        report.add("injection_count_invalid", "irrelevant record injection must add 10%-30%", actual=ratio)
    if attack_type == "near_match_record_injection" and ratio > 0.3:
        report.add("injection_count_invalid", "near-match injection may not exceed 30%", actual=ratio)
    _validate_declared_injected_rows(payload, file_name, None, len(left), len(right), report)
    if attack_type == "aggregate_row_injection":
        _validate_csv_aggregate_specs(left, right, file_name, payload, report)
    report.changed = added > 0


def _validate_csv_reshape(left: List[List[str]], right: List[List[str]], file_name: str, payload: Dict[str, Any], report: IntegrityReport) -> None:
    if not left or not right or left == right:
        report.add("reshape_not_changed", "CSV shape/content did not change", file=file_name)
        return
    report.changed = True
    spec = _reshape_spec(payload, file_name)
    if not spec:
        report.add("reshape_spec_missing", "wide/long reshape must declare one reshape_specs item per file", file=file_name)
        return
    if len(set(map(tuple, right[1:]))) != len(right) - 1:
        report.add("reshape_duplicate_rows", "reshaped CSV contains duplicate rows", file=file_name)
    direction = str(spec.get("direction", ""))
    id_vars = [str(item) for item in spec.get("id_vars") or []]
    value_vars = [str(item) for item in spec.get("value_vars") or []]
    dimension = str(spec.get("dimension_column", ""))
    value_column = str(spec.get("value_column", ""))
    if not id_vars or not value_vars or not dimension or not value_column:
        report.add("reshape_spec_incomplete", "reshape spec requires direction, id_vars, value_vars, dimension_column and value_column", file=file_name)
    elif direction == "wide_to_long":
        _validate_wide_to_long_inverse(left, right, file_name, id_vars, value_vars, dimension, value_column, report)
    elif direction == "long_to_wide":
        _validate_long_to_wide_inverse(left, right, file_name, id_vars, value_vars, dimension, value_column, report)
    else:
        report.add("reshape_direction_invalid", "direction must be wide_to_long or long_to_wide", file=file_name, actual=direction)
    report.checks[f"{file_name}/shape"] = {"original": [len(left), len(left[0])], "final": [len(right), len(right[0])]}


def _validate_relational_package(workspace: Any, report: IntegrityReport, payload: Dict[str, Any]) -> None:
    original_names = workspace.original_file_names()
    final_names = workspace.final_file_names()
    if len(original_names) != 1 or len(final_names) not in {2, 3}:
        report.add("relational_file_count_invalid", "one CSV must become two or three CSV files", expected="1 -> 2 or 3", actual=f"{len(original_names)} -> {len(final_names)}")
        return
    if any(Path(name).suffix.lower() not in {".csv", ".tsv"} for name in original_names + final_names):
        report.add("relational_file_type_invalid", "relational decomposition only supports CSV/TSV")
        return
    params = (payload.get("transformation_record") or {}).get("parameters") or {}
    key = params.get("key") or params.get("join_key") or params.get("keys")
    keys = [str(item) for item in key] if isinstance(key, list) else ([str(key)] if key else [])
    if not keys:
        report.add("join_key_missing", "relational decomposition must declare a join key")
        return
    original_delimiter = "\t" if Path(original_names[0]).suffix.lower() == ".tsv" else ","
    original = _read_delimited(workspace.original / original_names[0], original_delimiter)
    if not original:
        report.add("invalid_relational_source", "original relation is empty", file=original_names[0])
        return
    tables = []
    non_key_owners: Dict[str, str] = {}
    for name in final_names:
        delimiter = "\t" if Path(name).suffix.lower() == ".tsv" else ","
        rows = _read_delimited(workspace.final / name, delimiter)
        if len(rows) < 2 or len(set(rows[0])) != len(rows[0]):
            report.add("invalid_relational_table", "output table is empty or has duplicate headers", file=name)
            continue
        missing_keys = [item for item in keys if item not in rows[0]]
        if missing_keys:
            report.add("join_key_missing_from_table", "every output table must contain all join keys", file=name, actual=missing_keys)
            continue
        for header in rows[0]:
            if header in keys:
                continue
            if header in non_key_owners:
                report.add("field_duplicated_across_tables", "non-key field appears in multiple output tables", expected=non_key_owners[header], actual=name)
            non_key_owners[header] = name
        tables.append((name, rows))
    if len(tables) == len(final_names):
        _validate_relational_inverse(original, tables, keys, report)
    report.changed = bool(final_names)
    report.checks["final_files"] = final_names


def _hierarchical_sheet_spec(payload: Dict[str, Any], file_name: str, sheet_name: str) -> Dict[str, Any]:
    params = (payload.get("transformation_record") or {}).get("parameters") or {}
    specs = params.get("sheet_specs")
    if isinstance(specs, list):
        for spec in specs:
            if not isinstance(spec, dict):
                continue
            if Path(str(spec.get("file", ""))).name == file_name and str(spec.get("sheet", "")) == sheet_name:
                return spec
    required = {"original_data_start_row", "final_data_start_row"}
    return params if required.issubset(params) else {}


def _reshape_spec(payload: Dict[str, Any], file_name: str) -> Dict[str, Any]:
    params = (payload.get("transformation_record") or {}).get("parameters") or {}
    specs = params.get("reshape_specs")
    if isinstance(specs, list):
        for spec in specs:
            if isinstance(spec, dict) and Path(str(spec.get("file", ""))).name == file_name:
                return spec
        return {}
    required = {"direction", "id_vars", "value_vars", "dimension_column", "value_column"}
    return params if required.issubset(params) else {}


def _validate_wide_to_long_inverse(
    left: List[List[str]],
    right: List[List[str]],
    file_name: str,
    id_vars: List[str],
    value_vars: List[str],
    dimension: str,
    value_column: str,
    report: IntegrityReport,
) -> None:
    left_header, right_header = left[0], right[0]
    if set(id_vars + value_vars) != set(left_header) or len(id_vars) + len(value_vars) != len(left_header):
        report.add("reshape_fields_incomplete", "id_vars and value_vars must partition every original field", file=file_name, expected=left_header, actual=id_vars + value_vars)
        return
    expected_right = id_vars + [dimension, value_column]
    if set(right_header) != set(expected_right) or len(right_header) != len(expected_right):
        report.add("reshape_output_schema_invalid", "long output schema does not match declared fields", file=file_name, expected=expected_right, actual=right_header)
        return
    left_indices = {name: left_header.index(name) for name in left_header}
    right_indices = {name: right_header.index(name) for name in right_header}
    source_keys = [tuple(row[left_indices[name]] for name in id_vars) for row in left[1:]]
    if len(set(source_keys)) != len(source_keys):
        report.add("reshape_source_key_not_unique", "wide-table id_vars are not unique; inverse pivot would require aggregation", file=file_name)
        return
    long_values: Dict[Tuple[str, ...], str] = {}
    for row in right[1:]:
        if len(row) != len(right_header):
            report.add("reshape_row_width_invalid", "long row width does not match header", file=file_name)
            return
        key = tuple(row[right_indices[name]] for name in id_vars) + (row[right_indices[dimension]],)
        if key in long_values:
            report.add("reshape_key_not_unique", "id_vars plus dimension are not unique", file=file_name, actual=key)
            return
        long_values[key] = row[right_indices[value_column]]
    expected_count = len(left[1:]) * len(value_vars)
    if len(long_values) != expected_count:
        report.add("reshape_row_count_invalid", "long table row count does not match exact melt cardinality", file=file_name, expected=expected_count, actual=len(long_values))
    for row in left[1:]:
        base = tuple(row[left_indices[name]] for name in id_vars)
        for field in value_vars:
            key = base + (field,)
            if long_values.get(key) != row[left_indices[field]]:
                report.add("reshape_inverse_mismatch", "inverse pivot cannot recover an original lexical value", file=file_name, expected=row[left_indices[field]], actual=long_values.get(key))
                return
    report.checks[f"{file_name}/inverse_verified"] = True


def _validate_long_to_wide_inverse(
    left: List[List[str]],
    right: List[List[str]],
    file_name: str,
    id_vars: List[str],
    value_vars: List[str],
    dimension: str,
    value_column: str,
    report: IntegrityReport,
) -> None:
    left_header, right_header = left[0], right[0]
    expected_left = id_vars + [dimension, value_column]
    if set(left_header) != set(expected_left) or len(left_header) != len(expected_left):
        report.add("reshape_source_schema_invalid", "long source schema does not match declared fields", file=file_name, expected=expected_left, actual=left_header)
        return
    expected_right = id_vars + value_vars
    if set(right_header) != set(expected_right) or len(right_header) != len(expected_right):
        report.add("reshape_output_schema_invalid", "wide output schema does not match declared fields", file=file_name, expected=expected_right, actual=right_header)
        return
    left_indices = {name: left_header.index(name) for name in left_header}
    right_indices = {name: right_header.index(name) for name in right_header}
    source: Dict[Tuple[str, ...], str] = {}
    for row in left[1:]:
        key = tuple(row[left_indices[name]] for name in id_vars) + (row[left_indices[dimension]],)
        if key in source:
            report.add("reshape_source_key_not_unique", "long-table id_vars plus dimension are not unique", file=file_name, actual=key)
            return
        source[key] = row[left_indices[value_column]]
    wide_keys = [tuple(row[right_indices[name]] for name in id_vars) for row in right[1:]]
    if len(set(wide_keys)) != len(wide_keys):
        report.add("reshape_key_not_unique", "wide-table id_vars are not unique", file=file_name)
        return
    for row in right[1:]:
        base = tuple(row[right_indices[name]] for name in id_vars)
        for field in value_vars:
            if source.get(base + (field,)) != row[right_indices[field]]:
                report.add("reshape_inverse_mismatch", "wide value does not recover the original long lexical value", file=file_name, expected=source.get(base + (field,)), actual=row[right_indices[field]])
                return
    if len(source) != len(right[1:]) * len(value_vars):
        report.add("reshape_row_count_invalid", "wide table does not cover every long key/dimension pair", file=file_name)
    report.checks[f"{file_name}/inverse_verified"] = True


def _validate_relational_inverse(
    original: List[List[str]],
    tables: List[Tuple[str, List[List[str]]]],
    keys: List[str],
    report: IntegrityReport,
) -> None:
    original_header = original[0]
    if len(set(original_header)) != len(original_header) or any(key not in original_header for key in keys):
        report.add("invalid_relational_source", "original relation has duplicate headers or lacks declared keys")
        return
    union_headers = set(keys)
    for _, rows in tables:
        union_headers.update(rows[0])
    if union_headers != set(original_header):
        report.add("relational_fields_not_lossless", "output tables must preserve every original field and add none", expected=original_header, actual=sorted(union_headers))
        return
    base_name, base_rows = max(tables, key=lambda item: len(item[1]))
    base_header = base_rows[0]
    base_indices = [base_header.index(key) for key in keys]
    lookups: List[Tuple[str, Dict[Tuple[str, ...], Dict[str, str]]]] = []
    for name, rows in tables:
        if name == base_name:
            continue
        header = rows[0]
        key_indices = [header.index(key) for key in keys]
        lookup: Dict[Tuple[str, ...], Dict[str, str]] = {}
        for row in rows[1:]:
            if len(row) != len(header):
                report.add("relational_row_width_invalid", "output row width does not match header", file=name)
                return
            key = tuple(row[index] for index in key_indices)
            if any(value == "" for value in key):
                report.add("null_join_key", "join keys must be non-empty", file=name, actual=key)
                return
            if key in lookup:
                report.add("unsafe_join_cardinality", "non-base relation has duplicate keys and would cause many-to-many expansion", file=name, actual=key)
                return
            lookup[key] = dict(zip(header, row))
        lookups.append((name, lookup))
    reconstructed = []
    base_key_set = set()
    for row in base_rows[1:]:
        if len(row) != len(base_header):
            report.add("relational_row_width_invalid", "base output row width does not match header", file=base_name)
            return
        key = tuple(row[index] for index in base_indices)
        if any(value == "" for value in key):
            report.add("null_join_key", "join keys must be non-empty", file=base_name, actual=key)
            return
        base_key_set.add(key)
        merged = dict(zip(base_header, row))
        for name, lookup in lookups:
            match = lookup.get(key)
            if match is None:
                report.add("join_record_missing", "a base record has no matching related record", file=name, actual=key)
                return
            merged.update(match)
        reconstructed.append(tuple(merged.get(header) for header in original_header))
    for name, lookup in lookups:
        if set(lookup) != base_key_set:
            report.add("orphan_join_record", "related table key set differs from base relation", file=name)
            return
    expected = Counter(tuple(row) for row in original[1:])
    actual = Counter(reconstructed)
    if expected != actual:
        report.add("relational_inverse_mismatch", "joining output tables does not recover the original lexical rows", expected=f"{sum(expected.values())} rows", actual=f"{sum(actual.values())} rows")
        return
    report.checks["join_key"] = keys
    report.checks["reconstructed_rows"] = len(reconstructed)
    report.checks["inverse_verified"] = True


def _read_delimited(path: Path, delimiter: str) -> List[List[str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as stream:
        return list(csv.reader(stream, delimiter=delimiter))


def _container_file_metadata_text(path: Path) -> str:
    suffix = path.suffix.lower()
    try:
        if suffix in {".xlsx", ".xlsm"}:
            from openpyxl import load_workbook

            workbook = load_workbook(path, data_only=False, read_only=True, keep_vba=suffix == ".xlsm")
            try:
                values = [path.name]
                for sheet in workbook.worksheets[:3]:
                    values.append(sheet.title)
                    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 8), values_only=True):
                        values.extend(str(value) for value in row[:8] if value is not None)
                return " ".join(values)
            finally:
                workbook.close()
        if suffix in {".csv", ".tsv"}:
            delimiter = "\t" if suffix == ".tsv" else ","
            rows = _read_delimited(path, delimiter)[:8]
            return " ".join([path.name] + [value for row in rows for value in row[:8]])
    except Exception:
        return path.name
    return path.name


def _load_original_workbook_cached(path: Path, keep_vba: bool) -> Any:
    from openpyxl import load_workbook

    stat = path.stat()
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        digest.update(stream.read(65536))
        if stat.st_size > 65536:
            stream.seek(max(0, stat.st_size - 65536))
            digest.update(stream.read(65536))
    key = (path.name, stat.st_size, stat.st_mtime_ns, digest.hexdigest(), keep_vba)
    with _ORIGINAL_WORKBOOK_CACHE_LOCK:
        cached = _ORIGINAL_WORKBOOK_CACHE.get(key)
        if cached is not None:
            _ORIGINAL_WORKBOOK_CACHE.move_to_end(key)
            return cached
        workbook = load_workbook(path, data_only=False, read_only=False, keep_vba=keep_vba)
        _ORIGINAL_WORKBOOK_CACHE[key] = workbook
        while len(_ORIGINAL_WORKBOOK_CACHE) > _ORIGINAL_WORKBOOK_CACHE_SIZE:
            _, evicted = _ORIGINAL_WORKBOOK_CACHE.popitem(last=False)
            evicted.close()
        return workbook


def _data_token(cell: Any) -> Tuple[Any, ...]:
    data_type = "null" if cell.value is None else cell.data_type
    return _value_token(cell.value), data_type, cell.number_format


def _header_token(cell: Any) -> Tuple[Any, ...]:
    return _data_token(cell), tuple(cell._style) if cell.has_style else ()


def _header_nonvalue_token(cell: Any) -> Tuple[Any, ...]:
    return type(cell.value).__name__, cell.data_type, cell.number_format, tuple(cell._style) if cell.has_style else ()


def _column_token(cell: Any, is_header: bool) -> Tuple[Any, ...]:
    return _header_token(cell) if is_header else _data_token(cell)


def _value_token(value: Any) -> Tuple[str, str]:
    return type(value).__name__, repr(value)


def _formula_token(cell: Any) -> Tuple[Any, ...]:
    return _value_token(cell.value), cell.number_format


def _sheet_matrix_token(sheet: Any) -> Tuple[Any, ...]:
    return tuple(tuple(_header_token(cell) if row == 1 else _data_token(cell) for cell in cells) for row, cells in enumerate(sheet.iter_rows(), 1))


def _canonical_workbook_rows(workbook: Any) -> Counter:
    rows: Counter = Counter()
    for sheet in workbook.worksheets:
        headers = [cell.value for cell in sheet[1]]
        if len(set(headers)) != len(headers):
            continue
        for values in sheet.iter_rows(min_row=2):
            record = tuple(sorted((str(headers[index]), _data_token(cell)) for index, cell in enumerate(values)))
            rows[record] += 1
    return rows


def _header_styles_by_name(workbook: Any) -> Dict[str, set]:
    result: Dict[str, set] = {}
    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            if cell.value is None:
                continue
            result.setdefault(str(cell.value), set()).add(tuple(cell._style) if cell.has_style else ())
    return result


def _filter_token(sheet: Any) -> Tuple[Any, ...]:
    filters = []
    for item in getattr(sheet.auto_filter, "filterColumn", []) or []:
        filters.append(repr(item))
    return sheet.auto_filter.ref, tuple(filters)


def _table_tokens(sheet: Any) -> Tuple[Tuple[str, str], ...]:
    return tuple(sorted((name, table.ref) for name, table in sheet.tables.items()))


def _hidden_rows(sheet: Any) -> Tuple[int, ...]:
    return tuple(sorted(index for index, dim in sheet.row_dimensions.items() if dim.hidden))


def _hidden_columns(sheet: Any) -> Tuple[str, ...]:
    return tuple(sorted(index for index, dim in sheet.column_dimensions.items() if dim.hidden))


def _semantic_contract(payload: Dict[str, Any]) -> Dict[str, Any]:
    record = payload.get("transformation_record") or {}
    contract = record.get("semantic_contract")
    return contract if isinstance(contract, dict) else {}


def _direct_mapping(payload: Dict[str, Any]) -> Dict[str, str]:
    record = payload.get("transformation_record") or {}
    mapping = record.get("mapping")
    if not isinstance(mapping, dict):
        return {}
    return {str(key): str(value) for key, value in mapping.items() if isinstance(key, str) and isinstance(value, str)}


def _selection_fields(selection: Dict[str, Any]) -> set[str]:
    target = selection.get("target") or {}
    fields = target.get("fields") if isinstance(target, dict) else []
    return {str(item) for item in fields or [] if str(item).strip()}


def _has_text(value: Any) -> bool:
    return isinstance(value, str) and bool(value.strip())


def _validate_l1_semantic_contract(
    attack_type: str,
    payload: Dict[str, Any],
    selection: Dict[str, Any],
    report: IntegrityReport,
) -> None:
    contract = _semantic_contract(payload)
    if not contract:
        report.add("semantic_contract_missing", "L1 attacks require transformation_record.semantic_contract")
        return
    selected = _selection_fields(selection)
    if attack_type == "header_synonym_substitution":
        mapping = _direct_mapping(payload)
        if not 1 <= len(mapping) <= 3:
            report.add("synonym_mapping_count_invalid", "header synonym attack must rename one to three fields", expected="1..3", actual=len(mapping))
        bindings = contract.get("field_bindings")
        audits = contract.get("synonym_audit")
        if not isinstance(bindings, list) or len(bindings) != len(mapping):
            report.add("field_bindings_invalid", "field_bindings must contain exactly one item per header mapping", expected=len(mapping), actual=len(bindings) if isinstance(bindings, list) else None)
            bindings = []
        if not isinstance(audits, list) or len(audits) != len(mapping):
            report.add("synonym_audit_invalid", "synonym_audit must contain exactly one item per header mapping", expected=len(mapping), actual=len(audits) if isinstance(audits, list) else None)
            audits = []
        binding_pairs = set()
        for binding in bindings:
            if not isinstance(binding, dict):
                report.add("field_binding_invalid", "each field binding must be an object")
                continue
            old = str(binding.get("target_field", "")).strip()
            new = str(binding.get("transformed_field", "")).strip()
            binding_pairs.add((old, new))
            if not _has_text(binding.get("question_concept")) or not _has_text(binding.get("question_evidence")):
                report.add("field_binding_evidence_missing", "field binding requires question_concept and question_evidence", expected=old, actual=new)
            if selected and old not in selected and not _has_text(binding.get("selection_deviation_reason")):
                report.add("selection_target_drift", "renamed field is outside selection target without a deviation reason", expected=sorted(selected), actual=old)
        if binding_pairs != set(mapping.items()):
            report.add("field_binding_mapping_mismatch", "field bindings must exactly match transformation mapping", expected=sorted(mapping.items()), actual=sorted(binding_pairs))
        audit_pairs = set()
        required_true = ("same_concept", "same_metric_scope", "same_granularity", "same_time_basis", "same_unit")
        for audit in audits:
            if not isinstance(audit, dict):
                report.add("synonym_audit_item_invalid", "each synonym audit must be an object")
                continue
            old = str(audit.get("old_header", "")).strip()
            new = str(audit.get("new_header", "")).strip()
            audit_pairs.add((old, new))
            if not all(audit.get(field) is True for field in required_true) or audit.get("can_coexist_as_distinct") is not False:
                report.add("synonym_audit_failed", "constructor did not establish strict semantic equivalence", expected=old, actual=new)
            if not _has_text(audit.get("evidence")):
                report.add("synonym_evidence_missing", "synonym audit requires concrete evidence", expected=old, actual=new)
        if audit_pairs != set(mapping.items()):
            report.add("synonym_audit_mapping_mismatch", "synonym audits must exactly match transformation mapping", expected=sorted(mapping.items()), actual=sorted(audit_pairs))
    else:
        required = (
            "target_concept",
            "target_field",
            "distractor_field",
            "question_binding_quote",
            "exclusion_evidence",
            "counterfactual_method",
            "counterfactual_answer",
        )
        missing = [field for field in required if not _has_text(contract.get(field))]
        if missing:
            report.add("distractor_contract_incomplete", "distractor semantic contract is incomplete", actual=missing)
        target_field = str(contract.get("target_field", "")).strip()
        distractor_field = str(contract.get("distractor_field", "")).strip()
        if target_field == distractor_field:
            report.add("distractor_matches_target", "distractor field name must differ from target field")
        if selected and target_field not in selected and not _has_text(contract.get("selection_deviation_reason")):
            report.add("selection_target_drift", "distractor target is outside selection without a deviation reason", expected=sorted(selected), actual=target_field)


def _value_family(value: Any, data_type: str = "") -> str:
    if value in (None, ""):
        return "empty"
    if data_type == "f":
        return "formula"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (datetime, date)):
        return "date"
    if _is_number(value):
        return "number"
    return "text"


def _validate_distractor_values(sheet: Any, target_col: int, distractor_col: int, file_name: str, report: IntegrityReport) -> None:
    target = [sheet.cell(row, target_col) for row in range(2, sheet.max_row + 1)]
    distractor = [sheet.cell(row, distractor_col) for row in range(2, sheet.max_row + 1)]
    target_nonempty = [cell for cell in target if cell.value not in (None, "")]
    distractor_nonempty = [cell for cell in distractor if cell.value not in (None, "")]
    coverage = len(distractor_nonempty) / max(1, len(target_nonempty))
    report.checks[f"{file_name}/{sheet.title}/distractor_nonempty_coverage"] = round(coverage, 6)
    if coverage < 0.5:
        report.add("distractor_coverage_too_low", "distractor non-empty coverage must be at least 50% of target", file=file_name, sheet=sheet.title, actual=coverage)
    target_families = {_value_family(cell.value, cell.data_type) for cell in target_nonempty}
    distractor_families = {_value_family(cell.value, cell.data_type) for cell in distractor_nonempty}
    if target_families and distractor_families and target_families.isdisjoint(distractor_families):
        report.add("distractor_type_incompatible", "target and distractor have incompatible value types", file=file_name, sheet=sheet.title, expected=sorted(target_families), actual=sorted(distractor_families))
    comparable = [(a.value, b.value) for a, b in zip(target, distractor) if a.value not in (None, "") and b.value not in (None, "")]
    difference_ratio = sum(a != b for a, b in comparable) / max(1, len(comparable))
    report.checks[f"{file_name}/{sheet.title}/distractor_difference_ratio"] = round(difference_ratio, 6)
    if difference_ratio < 0.3:
        report.add("distractor_difference_too_low", "fewer than 30% of comparable distractor values differ from target", file=file_name, sheet=sheet.title, actual=difference_ratio)


def _lexical_family(value: str) -> str:
    text = str(value).strip()
    if not text:
        return "empty"
    try:
        float(text.replace(",", ""))
        return "number"
    except ValueError:
        pass
    if re.fullmatch(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}", text):
        return "date"
    return "text"


def _validate_lexical_distractor_values(target: List[str], distractor: List[str], file_name: str, report: IntegrityReport) -> None:
    target_nonempty = [value for value in target if str(value).strip()]
    distractor_nonempty = [value for value in distractor if str(value).strip()]
    coverage = len(distractor_nonempty) / max(1, len(target_nonempty))
    report.checks[f"{file_name}/distractor_nonempty_coverage"] = round(coverage, 6)
    if coverage < 0.5:
        report.add("distractor_coverage_too_low", "CSV distractor non-empty coverage must be at least 50%", file=file_name, actual=coverage)
    target_families = {_lexical_family(value) for value in target_nonempty}
    distractor_families = {_lexical_family(value) for value in distractor_nonempty}
    if target_families and distractor_families and target_families.isdisjoint(distractor_families):
        report.add("distractor_type_incompatible", "CSV target and distractor have incompatible lexical types", file=file_name, expected=sorted(target_families), actual=sorted(distractor_families))
    comparable = [(a, b) for a, b in zip(target, distractor) if str(a).strip() and str(b).strip()]
    ratio = sum(a != b for a, b in comparable) / max(1, len(comparable))
    report.checks[f"{file_name}/distractor_difference_ratio"] = round(ratio, 6)
    if ratio < 0.3:
        report.add("distractor_difference_too_low", "fewer than 30% of CSV distractor values differ", file=file_name, actual=ratio)


def _declared_string_mappings(payload: Dict[str, Any]) -> Dict[str, str]:
    record = payload.get("transformation_record") or {}
    candidates: List[Dict[str, str]] = []

    def visit(value: Any) -> None:
        if isinstance(value, dict):
            pairs = {str(key): str(item) for key, item in value.items() if isinstance(key, str) and isinstance(item, str)}
            if pairs:
                candidates.append(pairs)
            for item in value.values():
                visit(item)
        elif isinstance(value, list):
            for item in value:
                visit(item)

    visit(record.get("mapping"))
    visit(record.get("parameters"))
    result: Dict[str, str] = {}
    for candidate in sorted(candidates, key=len):
        result.update(candidate)
    return result


def _equivalent_value(left: Any, right: Any) -> bool:
    if left == right and type(left) is type(right):
        return True
    left_date = _parse_unambiguous_date(left)
    right_date = _parse_unambiguous_date(right)
    if left_date is not None or right_date is not None:
        return left_date is not None and right_date is not None and left_date == right_date
    left_number = _parse_equivalent_number(left)
    right_number = _parse_equivalent_number(right)
    if left_number is not None or right_number is not None:
        return left_number is not None and right_number is not None and math.isclose(
            left_number, right_number, rel_tol=1e-12, abs_tol=1e-12
        )
    if isinstance(left, bool) and isinstance(right, str):
        truth = right.strip().lower()
        return (left and truth in {"是", "yes", "true", "y"}) or (not left and truth in {"否", "no", "false", "n"})
    if isinstance(right, bool) and isinstance(left, str):
        return _equivalent_value(right, left)
    return str(left).strip() == str(right).strip()


def _parse_unambiguous_date(value: Any) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if not isinstance(value, str):
        return None
    text = value.strip()
    for pattern in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(text, pattern)
        except ValueError:
            continue
    return None


def _parse_equivalent_number(value: Any) -> Optional[float]:
    if isinstance(value, bool):
        return None
    if _is_number(value):
        return float(value)
    if not isinstance(value, str):
        return None
    text = value.strip().replace(",", "")
    percent = text.endswith("%")
    if percent:
        text = text[:-1].strip()
    try:
        parsed = float(text)
    except ValueError:
        return None
    return parsed / 100 if percent else parsed


def _is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _short(value: Any, limit: int = 300) -> Optional[str]:
    if value is None:
        return None
    text = repr(value)
    return text if len(text) <= limit else text[:limit] + "..."
