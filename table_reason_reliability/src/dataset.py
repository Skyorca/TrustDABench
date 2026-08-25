from __future__ import annotations

import json
import math
import re
import time
import zipfile
from datetime import date, datetime, time as datetime_time
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import pandas as pd


# Profile 只用于帮助模型判断攻击是否适用，不需要读取整张表。
PROFILE_SAMPLE_ROWS = 8
HEADER_SCAN_ROWS = 3
MAX_PROFILE_COLUMNS = 200
MAX_PROFILE_SHEETS = 20

# 只对较小的 worksheet XML 检查合并单元格。
# 超过该大小直接跳过，避免再次长时间占用 CPU。
MAX_STRUCTURE_XML_BYTES = 4 * 1024 * 1024

# CSV 小于该大小时统计完整行数；大文件只估算行数。
MAX_CSV_COUNT_BYTES = 64 * 1024 * 1024


def load_jsonl(path: Path) -> List[Dict[str, Any]]:
    """读取 UTF-8/UTF-8-BOM JSONL 文件。"""
    path = Path(path)
    rows: List[Dict[str, Any]] = []

    with path.open("r", encoding="utf-8-sig") as file:
        for line_number, line in enumerate(file, start=1):
            text = line.strip()
            if not text:
                continue

            try:
                payload = json.loads(text)
            except json.JSONDecodeError as exc:
                raise ValueError(
                    f"Invalid JSONL at {path}, line {line_number}: {exc}"
                ) from exc

            if not isinstance(payload, dict):
                raise ValueError(
                    f"JSONL row must be an object: {path}, line {line_number}"
                )

            rows.append(payload)

    return rows


def rows_to_limit(
    rows: Sequence[Dict[str, Any]],
    limit: Optional[int],
) -> List[Dict[str, Any]]:
    """根据 limit 截取样本。"""
    result = list(rows)

    if limit is None:
        return result

    limit_value = int(limit)
    if limit_value < 0:
        return result

    return result[:limit_value]


def compact_json(value: Any) -> str:
    """供 Prompt 使用的稳定 JSON 序列化。"""
    return json.dumps(
        value,
        ensure_ascii=False,
        indent=2,
        default=_json_safe,
    )


def split_input_files(value: Any) -> List[str]:
    """解析 input_file，支持列表、换行和中英文分号。"""
    if value is None:
        return []

    if isinstance(value, (list, tuple, set)):
        raw_values: Iterable[Any] = value
    else:
        raw_values = [value]

    names: List[str] = []
    seen = set()

    for raw_value in raw_values:
        if raw_value is None:
            continue

        text = str(raw_value).strip()
        if not text:
            continue

        # 文件名可能包含逗号，所以不能按逗号分割。
        parts = re.split(r"[\r\n;；]+", text)

        for part in parts:
            name = part.strip().strip('"').strip("'")
            if not name:
                continue

            normalized = name.replace("\\", "/")
            if normalized in seen:
                continue

            seen.add(normalized)
            names.append(name)

    return names


def resolve_input_files(
    row: Dict[str, Any],
    data_root: Path,
) -> List[Path]:
    """
    将样本 input_file 映射为本地文件路径。

    优先顺序：
    1. 绝对路径；
    2. data_root/<sample_id>/<filename>；
    3. data_root/<filename>。

    即使文件不存在，也保留最可能的路径，让 profile 明确报告缺失。
    """
    data_root = Path(data_root)
    sample_id = str(
        row.get("source_id")
        or row.get("id")
        or row.get("sample_id")
        or ""
    ).strip()

    file_names = split_input_files(
        row.get("input_file")
        or row.get("input_files")
        or row.get("files")
    )

    resolved: List[Path] = []
    seen = set()

    for file_name in file_names:
        raw_path = Path(file_name)

        if raw_path.is_absolute():
            candidates = [raw_path]
        else:
            candidates = []

            if sample_id:
                candidates.append(data_root / sample_id / raw_path)

            candidates.append(data_root / raw_path)

        selected = next(
            (candidate for candidate in candidates if candidate.exists()),
            candidates[0],
        )

        key = str(selected.resolve(strict=False))
        if key in seen:
            continue

        seen.add(key)
        resolved.append(selected)

    return resolved


def profile_files(
    files: Sequence[Path],
    virtual_root: str = "/mnt/data",
) -> Dict[str, Any]:
    """生成输入文件概况；任何单文件失败都不会阻断整个样本。"""
    paths = [Path(path) for path in files]
    profiles: List[Dict[str, Any]] = []

    for path in paths:
        started = time.monotonic()

        try:
            profile = profile_file(
                path,
                virtual_root=virtual_root,
            )
        except Exception as exc:
            profile = {
                "name": path.name,
                "virtual_path": (
                    f"{virtual_root.rstrip('/')}/{path.name}"
                ),
                "exists": path.exists(),
                "size_bytes": (
                    path.stat().st_size
                    if path.exists() and path.is_file()
                    else None
                ),
                "suffix": path.suffix.lower(),
                "file_type": "profile_error",
                "profile_error": f"{type(exc).__name__}: {exc}",
                "sheets": [],
            }

        profile["profile_elapsed_sec"] = round(
            time.monotonic() - started,
            4,
        )
        profiles.append(profile)

    return {
        "file_count": len(paths),
        "existing_file_count": sum(
            1 for path in paths if path.exists() and path.is_file()
        ),
        "all_files_exist": bool(paths) and all(
            path.exists() and path.is_file()
            for path in paths
        ),
        "files": profiles,
    }


def profile_file(
        path: Path,
        virtual_root: str = "/mnt/data",
    ) -> Dict[str, Any]:
    """按文件类型生成轻量 profile。"""
    path = Path(path)

    base: Dict[str, Any] = {
        "name": path.name,
        "virtual_path": (
            f"{virtual_root.rstrip('/')}/{path.name}"
        ),
        "exists": path.exists(),
        "size_bytes": (
            path.stat().st_size
            if path.exists() and path.is_file()
            else None
        ),
        "suffix": path.suffix.lower(),
    }

    if not path.exists() or not path.is_file():
        base.update(
            {
                "file_type": "missing",
                "profile_error": "file does not exist",
                "sheets": [],
            }
        )
        return base

    suffix = path.suffix.lower()

    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        base.update(_profile_modern_excel(path))
        return base

    if suffix == ".xls":
        base.update(_profile_legacy_excel(path))
        return base

    if suffix in {".csv", ".tsv", ".txt"}:
        base.update(_profile_delimited_file(path))
        return base

    if suffix in {".jsonl", ".ndjson"}:
        base.update(_profile_jsonl_file(path))
        return base

    if suffix == ".json":
        base.update(_profile_json_file(path))
        return base

    base.update(
        {
            "file_type": "unsupported",
            "profile_error": (
                f"unsupported file type: {suffix or '<no suffix>'}"
            ),
            "sheets": [],
        }
    )
    return base


def _profile_modern_excel(path: Path) -> Dict[str, Any]:
    """
    轻量读取 xlsx/xlsm。

    关键点：
    - read_only=True；
    - data_only=False，避免依赖公式缓存；
    - 最多读取前若干行、前200列；
    - 不调用 pd.read_excel 全量加载工作表。
    """
    from openpyxl import load_workbook

    workbook = load_workbook(
        filename=path,
        read_only=True,
        data_only=False,
        keep_links=False,
    )

    try:
        all_sheet_names = list(workbook.sheetnames)
        selected_sheet_names = all_sheet_names[:MAX_PROFILE_SHEETS]

        sheets: List[Dict[str, Any]] = []

        for sheet_name in selected_sheet_names:
            worksheet = workbook[sheet_name]
            sheets.append(_profile_openpyxl_sheet(worksheet))

        structure = _scan_xlsx_structure(
            path=path,
            sheet_names=all_sheet_names,
            sheet_profiles=sheets,
        )

        return {
            "file_type": "excel",
            "sheet_count": len(all_sheet_names),
            "profiled_sheet_count": len(selected_sheet_names),
            "sheet_profile_truncated": (
                len(all_sheet_names) > len(selected_sheet_names)
            ),
            "sheet_names": all_sheet_names,
            "sheets": sheets,
            "excel_structure": structure,
        }
    finally:
        workbook.close()


def _profile_openpyxl_sheet(worksheet: Any) -> Dict[str, Any]:
    max_row = int(worksheet.max_row or 0)
    max_column = int(worksheet.max_column or 0)

    read_column_count = min(max_column, MAX_PROFILE_COLUMNS)
    read_row_count = min(
        max_row,
        HEADER_SCAN_ROWS + PROFILE_SAMPLE_ROWS + 2,
    )

    preview_rows: List[List[Any]] = []

    if read_row_count > 0 and read_column_count > 0:
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=read_row_count,
            min_col=1,
            max_col=read_column_count,
            values_only=True,
        ):
            preview_rows.append(
                [_json_safe(value) for value in row]
            )

    header_index = _choose_header_row(preview_rows)
    raw_columns = (
        preview_rows[header_index]
        if preview_rows
        else []
    )

    columns, duplicate_base_names = _normalize_column_names(raw_columns)

    data_rows = preview_rows[
        header_index + 1:
        header_index + 1 + PROFILE_SAMPLE_ROWS
    ]

    sample_rows: List[Dict[str, Any]] = []

    for row in data_rows:
        record: Dict[str, Any] = {}

        for column_index, column_name in enumerate(columns):
            value = (
                row[column_index]
                if column_index < len(row)
                else None
            )
            record[column_name] = _json_safe(value)

        sample_rows.append(record)

    dtypes: Dict[str, str] = {}

    for column_index, column_name in enumerate(columns):
        values = []

        for row in data_rows:
            if column_index < len(row):
                values.append(row[column_index])

        dtypes[column_name] = _infer_dtype(values)

    unnamed_columns = [
        name for name in columns
        if name.lower().startswith("unnamed:")
    ]

    possible_multirow_header = (
        header_index > 0
        or _looks_like_multirow_header(
            preview_rows[:header_index + 1]
        )
    )

    return {
        "name": str(worksheet.title),
        "max_row": max_row,
        "max_column": max_column,
        "n_rows": max(0, max_row - header_index - 1),
        "column_count": max_column,
        "profiled_column_count": read_column_count,
        "column_profile_truncated": (
            max_column > read_column_count
        ),
        "header_row_number": header_index + 1,
        "columns": columns,
        "raw_columns": [
            _json_safe(value) for value in raw_columns
        ],
        "dtypes": dtypes,
        "sample_rows": sample_rows,
        "header_preview": preview_rows[:HEADER_SCAN_ROWS],
        "unnamed_column_count": len(unnamed_columns),
        "unnamed_columns": unnamed_columns,
        "duplicate_like_columns": [
            name for name in columns
            if re.search(r"\.\d+$", name)
        ],
        "duplicate_base_names": duplicate_base_names,
        "possible_multirow_header": possible_multirow_header,
        "possible_duplicate_header": bool(duplicate_base_names),
        "dimension_suspect": (
            max_row >= 1_000_000
            or max_column >= 10_000
        ),
    }


def _profile_legacy_excel(path: Path) -> Dict[str, Any]:
    """
    旧版 .xls 使用 pandas，但每个Sheet只读取少量行。
    不使用 sheet_name=None 全量读取。
    """
    excel_file = pd.ExcelFile(path)
    all_sheet_names = list(excel_file.sheet_names)
    selected_sheet_names = all_sheet_names[:MAX_PROFILE_SHEETS]

    sheets: List[Dict[str, Any]] = []

    for sheet_name in selected_sheet_names:
        frame = pd.read_excel(
            path,
            sheet_name=sheet_name,
            nrows=HEADER_SCAN_ROWS + PROFILE_SAMPLE_ROWS,
            header=None,
            dtype=object,
        )

        sheets.append(
            _profile_preview_frame(
                frame=frame,
                sheet_name=sheet_name,
                n_rows=None,
            )
        )

    return {
        "file_type": "legacy_excel",
        "sheet_count": len(all_sheet_names),
        "profiled_sheet_count": len(selected_sheet_names),
        "sheet_profile_truncated": (
            len(all_sheet_names) > len(selected_sheet_names)
        ),
        "sheet_names": all_sheet_names,
        "sheets": sheets,
        "excel_structure": {
            "sheet_count": len(all_sheet_names),
            "has_multiple_sheets": len(all_sheet_names) > 1,
            "structure_scan_truncated": True,
            "complex_structure_candidate": len(all_sheet_names) > 1,
        },
    }


def _profile_delimited_file(path: Path) -> Dict[str, Any]:
    encoding = _detect_text_encoding(path)
    suffix = path.suffix.lower()

    separator: Optional[str]
    if suffix == ".tsv":
        separator = "\t"
    else:
        separator = None

    read_kwargs: Dict[str, Any] = {
        "nrows": HEADER_SCAN_ROWS + PROFILE_SAMPLE_ROWS,
        "header": None,
        "dtype": object,
        "encoding": encoding,
    }

    if separator is None:
        read_kwargs.update(
            {
                "sep": None,
                "engine": "python",
            }
        )
    else:
        read_kwargs["sep"] = separator

    frame = pd.read_csv(path, **read_kwargs)

    n_rows, row_count_estimated = _count_or_estimate_text_rows(path)

    sheet = _profile_preview_frame(
        frame=frame,
        sheet_name="CSV",
        n_rows=n_rows,
    )

    return {
        "file_type": "delimited_text",
        "encoding": encoding,
        "row_count_estimated": row_count_estimated,
        "sheets": [sheet],
        "excel_structure": None,
    }


def _profile_preview_frame(
    frame: pd.DataFrame,
    sheet_name: str,
    n_rows: Optional[int],
) -> Dict[str, Any]:
    preview_rows = [
        [_json_safe(value) for value in row]
        for row in frame.itertuples(index=False, name=None)
    ]

    header_index = _choose_header_row(preview_rows)
    raw_columns = (
        preview_rows[header_index]
        if preview_rows
        else []
    )

    columns, duplicate_base_names = _normalize_column_names(raw_columns)

    data_rows = preview_rows[
        header_index + 1:
        header_index + 1 + PROFILE_SAMPLE_ROWS
    ]

    sample_rows = []

    for row in data_rows:
        sample_rows.append(
            {
                column_name: (
                    _json_safe(row[index])
                    if index < len(row)
                    else None
                )
                for index, column_name in enumerate(columns)
            }
        )

    dtypes = {
        column_name: _infer_dtype(
            [
                row[index]
                for row in data_rows
                if index < len(row)
            ]
        )
        for index, column_name in enumerate(columns)
    }

    unnamed_columns = [
        name for name in columns
        if name.lower().startswith("unnamed:")
    ]

    return {
        "name": sheet_name,
        "n_rows": n_rows,
        "column_count": len(columns),
        "profiled_column_count": len(columns),
        "column_profile_truncated": False,
        "header_row_number": header_index + 1,
        "columns": columns,
        "raw_columns": raw_columns,
        "dtypes": dtypes,
        "sample_rows": sample_rows,
        "header_preview": preview_rows[:HEADER_SCAN_ROWS],
        "unnamed_column_count": len(unnamed_columns),
        "unnamed_columns": unnamed_columns,
        "duplicate_like_columns": [
            name for name in columns
            if re.search(r"\.\d+$", name)
        ],
        "duplicate_base_names": duplicate_base_names,
        "possible_multirow_header": (
            header_index > 0
            or _looks_like_multirow_header(
                preview_rows[:header_index + 1]
            )
        ),
        "possible_duplicate_header": bool(duplicate_base_names),
    }


def _profile_jsonl_file(path: Path) -> Dict[str, Any]:
    rows = []
    keys = []
    seen_keys = set()
    line_count = 0

    with path.open("r", encoding="utf-8-sig") as file:
        for line in file:
            if not line.strip():
                continue

            line_count += 1

            if len(rows) >= PROFILE_SAMPLE_ROWS:
                continue

            payload = json.loads(line)

            if isinstance(payload, dict):
                safe_payload = {
                    str(key): _json_safe(value)
                    for key, value in payload.items()
                }
                rows.append(safe_payload)

                for key in safe_payload:
                    if key not in seen_keys:
                        seen_keys.add(key)
                        keys.append(key)
            else:
                rows.append({"value": _json_safe(payload)})

    return {
        "file_type": "jsonl",
        "n_rows": line_count,
        "columns": keys,
        "sample_rows": rows,
        "sheets": [],
    }


def _profile_json_file(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8-sig") as file:
        payload = json.load(file)

    if isinstance(payload, list):
        sample = payload[:PROFILE_SAMPLE_ROWS]
        n_rows = len(payload)
    else:
        sample = [payload]
        n_rows = 1

    return {
        "file_type": "json",
        "n_rows": n_rows,
        "sample_rows": [_json_safe(item) for item in sample],
        "sheets": [],
    }


def _choose_header_row(rows: Sequence[Sequence[Any]]) -> int:
    """
    只在前3行中选择最可能的表头。

    不继续扫描数据区，避免把字符串较多的数据行误认为表头。
    """
    if not rows:
        return 0

    candidate_count = min(HEADER_SCAN_ROWS, len(rows))
    best_index = 0
    best_score = float("-inf")

    for index in range(candidate_count):
        row = rows[index]

        nonempty = [
            value for value in row
            if not _is_empty(value)
        ]

        if not nonempty:
            score = -1000 + index
        else:
            string_count = sum(
                isinstance(value, str)
                and not str(value).startswith("=")
                for value in nonempty
            )
            formula_count = sum(
                isinstance(value, str)
                and str(value).startswith("=")
                for value in nonempty
            )
            numeric_count = sum(
                isinstance(value, (int, float))
                and not isinstance(value, bool)
                for value in nonempty
            )
            unique_count = len(
                {str(value).strip() for value in nonempty}
            )

            score = (
                string_count * 4
                + unique_count * 1.5
                + len(nonempty)
                - numeric_count * 2
                - formula_count * 3
            )

            # 只有一个非空单元格通常是标题，而不是字段行。
            if len(nonempty) == 1:
                score -= 8

        if score > best_score:
            best_score = score
            best_index = index

    return best_index


def _normalize_column_names(
    raw_columns: Sequence[Any],
) -> Tuple[List[str], List[str]]:
    columns: List[str] = []
    base_counts: Dict[str, int] = {}
    duplicate_base_names: List[str] = []

    for index, value in enumerate(raw_columns):
        if _is_empty(value):
            base_name = f"Unnamed: {index}"
        else:
            base_name = str(value).strip()

        count = base_counts.get(base_name, 0)
        base_counts[base_name] = count + 1

        if count == 0:
            final_name = base_name
        else:
            final_name = f"{base_name}.{count}"
            if base_name not in duplicate_base_names:
                duplicate_base_names.append(base_name)

        columns.append(final_name)

    return columns, duplicate_base_names


def _looks_like_multirow_header(
    rows: Sequence[Sequence[Any]],
) -> bool:
    if len(rows) <= 1:
        return False

    first_nonempty = sum(
        not _is_empty(value)
        for value in rows[0]
    )
    later_nonempty = max(
        sum(not _is_empty(value) for value in row)
        for row in rows[1:]
    )

    return (
        first_nonempty > 0
        and later_nonempty > first_nonempty
    )


def _infer_dtype(values: Sequence[Any]) -> str:
    types = set()

    for value in values:
        if _is_empty(value):
            continue

        if isinstance(value, str) and value.startswith("="):
            types.add("formula")
        elif isinstance(value, bool):
            types.add("bool")
        elif isinstance(value, int):
            types.add("int")
        elif isinstance(value, float):
            types.add("float")
        elif isinstance(value, (datetime, date, datetime_time)):
            types.add("datetime")
        elif isinstance(value, str):
            types.add("string")
        else:
            types.add(type(value).__name__)

    if not types:
        return "empty"

    if types <= {"int", "float"}:
        return "float" if "float" in types else "int"

    if len(types) == 1:
        return next(iter(types))

    return "mixed[" + ",".join(sorted(types)) + "]"


def _scan_xlsx_structure(
    path: Path,
    sheet_names: Sequence[str],
    sheet_profiles: Sequence[Dict[str, Any]],
) -> Dict[str, Any]:
    """
    从 xlsx ZIP 包轻量检查合并单元格。

    大 worksheet XML 不解压扫描，防止再次发生长时间卡住。
    """
    structure_sheets: List[Dict[str, Any]] = []
    structure_scan_truncated = False

    try:
        with zipfile.ZipFile(path) as archive:
            members = sorted(
                (
                    info for info in archive.infolist()
                    if re.fullmatch(
                        r"xl/worksheets/sheet\d+\.xml",
                        info.filename,
                    )
                ),
                key=lambda info: _sheet_member_number(info.filename),
            )

            for index, sheet_name in enumerate(
                sheet_names[:MAX_PROFILE_SHEETS]
            ):
                entry: Dict[str, Any] = {
                    "name": sheet_name,
                    "xml_member": None,
                    "merged_range_count": None,
                    "merged_ranges": [],
                    "merged_header_ranges": [],
                    "has_merged_header": False,
                    "structure_scan_truncated": False,
                }

                if index >= len(members):
                    entry["structure_scan_truncated"] = True
                    structure_scan_truncated = True
                    structure_sheets.append(entry)
                    continue

                info = members[index]
                entry["xml_member"] = info.filename
                entry["xml_size_bytes"] = info.file_size

                if info.file_size > MAX_STRUCTURE_XML_BYTES:
                    entry["structure_scan_truncated"] = True
                    structure_scan_truncated = True
                    structure_sheets.append(entry)
                    continue

                xml_bytes = archive.read(info.filename)

                ranges = [
                    match.decode("utf-8", errors="replace")
                    for match in re.findall(
                        rb"<mergeCell[^>]*\bref=\"([^\"]+)\"",
                        xml_bytes,
                    )
                ]

                header_ranges = [
                    cell_range
                    for cell_range in ranges
                    if _range_starts_in_header(cell_range)
                ]

                entry.update(
                    {
                        "merged_range_count": len(ranges),
                        "merged_ranges": ranges[:20],
                        "merged_header_ranges": header_ranges[:20],
                        "has_merged_header": bool(header_ranges),
                    }
                )
                structure_sheets.append(entry)

    except Exception as exc:
        structure_scan_truncated = True
        structure_sheets = []
        structure_error = f"{type(exc).__name__}: {exc}"
    else:
        structure_error = None

    has_multiple_sheets = len(sheet_names) > 1
    has_any_merged_headers = any(
        item.get("has_merged_header") is True
        for item in structure_sheets
    )
    has_possible_multirow_header = any(
        sheet.get("possible_multirow_header") is True
        for sheet in sheet_profiles
    )
    has_possible_duplicate_header = any(
        sheet.get("possible_duplicate_header") is True
        for sheet in sheet_profiles
    )

    return {
        "sheet_count": len(sheet_names),
        "has_multiple_sheets": has_multiple_sheets,
        "has_any_merged_headers": has_any_merged_headers,
        "has_possible_multirow_header": has_possible_multirow_header,
        "has_possible_duplicate_header": has_possible_duplicate_header,
        "complex_structure_candidate": (
            has_multiple_sheets
            or has_any_merged_headers
            or has_possible_multirow_header
            or has_possible_duplicate_header
        ),
        "structure_scan_truncated": structure_scan_truncated,
        "structure_error": structure_error,
        "sheets": structure_sheets,
    }


def _sheet_member_number(name: str) -> int:
    match = re.search(r"sheet(\d+)\.xml$", name)
    return int(match.group(1)) if match else 10**9


def _range_starts_in_header(cell_range: str) -> bool:
    start_cell = cell_range.split(":", 1)[0]
    match = re.search(r"(\d+)$", start_cell)

    if not match:
        return False

    return int(match.group(1)) <= 5


def _detect_text_encoding(path: Path) -> str:
    sample = path.read_bytes()[:256 * 1024]

    for encoding in (
        "utf-8-sig",
        "utf-8",
        "gb18030",
        "gbk",
        "big5",
    ):
        try:
            sample.decode(encoding)
            return encoding
        except UnicodeDecodeError:
            continue

    return "latin-1"


def _count_or_estimate_text_rows(
    path: Path,
) -> Tuple[Optional[int], bool]:
    file_size = path.stat().st_size

    if file_size <= MAX_CSV_COUNT_BYTES:
        newline_count = 0

        with path.open("rb") as file:
            for chunk in iter(
                lambda: file.read(1024 * 1024),
                b"",
            ):
                newline_count += chunk.count(b"\n")

        # 默认第一行为表头。
        return max(0, newline_count - 1), False

    with path.open("rb") as file:
        sample = file.read(1024 * 1024)

    sample_lines = max(1, sample.count(b"\n"))
    average_line_size = max(1, len(sample) / sample_lines)
    estimated_rows = max(
        0,
        int(file_size / average_line_size) - 1,
    )

    return estimated_rows, True


def _is_empty(value: Any) -> bool:
    if value is None:
        return True

    if isinstance(value, str):
        return not value.strip()

    try:
        result = pd.isna(value)
    except Exception:
        return False

    return bool(result) if isinstance(result, bool) else False


def _json_safe(value: Any) -> Any:
    if value is None:
        return None

    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return value

    if isinstance(value, (datetime, date, datetime_time)):
        return value.isoformat()

    if isinstance(value, Path):
        return str(value)

    if isinstance(value, dict):
        return {
            str(key): _json_safe(item)
            for key, item in value.items()
        }

    if isinstance(value, (list, tuple, set)):
        return [_json_safe(item) for item in value]

    # 处理 numpy 标量。
    if hasattr(value, "item"):
        try:
            return _json_safe(value.item())
        except Exception:
            pass

    try:
        result = pd.isna(value)
        if isinstance(result, bool) and result:
            return None
    except Exception:
        pass

    if isinstance(value, (str, int, float, bool)):
        return value

    return str(value)